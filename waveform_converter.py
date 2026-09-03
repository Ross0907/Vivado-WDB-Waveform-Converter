#!/usr/bin/env python3
"""Vivado/XSim waveform converter.

One self-contained tool for:
  * Vivado/XSim WDB or WCFG -> VCD, CSV, JSON, Excel
  * VCD -> CSV, JSON, Excel

The WDB decoder is pure Python. It does not launch Vivado or rerun simulations.
Its binary reader accepts the supported
`Xilinx WAVE DATABASE 01` + `Xilinx ISim DBG 006` format profile.
Unsupported or ambiguous mappings are rejected with diagnostics rather than
guessed.
"""
from __future__ import annotations

import argparse
from bisect import bisect_right
import csv
import fnmatch
import heapq
import json
import math
import queue
import re
import struct
import sys
import threading
import xml.etree.ElementTree as ET
import zlib
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterator

try:
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk
    HAS_TK = True
except ImportError:
    HAS_TK = False

APP_VERSION = "2.0.0"

WDB_MAGIC = b"Xilinx WAVE DATABASE 01\x00"
DBG_MAGIC = b"Xilinx ISim DBG 006\x00"
EVENT_MAGIC = b"WDB.Event\x00"
RTTI_TYPE_MAGIC = b"Xilinx ISim TYPE FILE 001\x00"
PAGE_SIZE = 0x4C0
ROOT0_REL = 0x30
PTR_ARRAY_OFF = 0x08
SIZE_ARRAY_OFF = 0x328
COUNT_OFF = 0x4B8
MAX_CHUNKS_PER_PAGE = 100
CHUNK_HEADER_SIZE = 20
EVENT_FIXED_SIZE = 16
DEFAULT_BANK_SIZE = 0x800

STRUCT_REC_SIZE = 36
OBJ_REC_SIZE = 44
ALLOC_REC_SIZE = 56


class DecodeError(RuntimeError):
    pass


def u32(b: bytes, off: int) -> int:
    if off < 0 or off + 4 > len(b):
        raise DecodeError(f"u32 outside file at 0x{off:X}")
    return struct.unpack_from("<I", b, off)[0]


def u64(b: bytes, off: int) -> int:
    if off < 0 or off + 8 > len(b):
        raise DecodeError(f"u64 outside file at 0x{off:X}")
    return struct.unpack_from("<Q", b, off)[0]


def is_power_of_two(x: int) -> bool:
    return x > 0 and (x & (x - 1)) == 0


class Logger:
    def __init__(self) -> None:
        self.lines: list[str] = []

    def add(self, text: str = "") -> None:
        self.lines.append(text)

    def text(self) -> str:
        return "\n".join(self.lines) + "\n"


@dataclass(frozen=True)
class DbgStructure:
    index: int
    name: str
    parent: int


@dataclass(frozen=True)
class RttiField:
    name: str
    type_index: int
    ranges: tuple[tuple[int, int, int], ...] = ()


@dataclass(frozen=True)
class RttiType:
    index: int
    size: int
    tag: int
    name: str
    raw: bytes
    kind: str = "sv_digital"
    language: str = "unknown"
    array_flavor: str | None = None
    element_type_index: int | None = None
    array_rank: int = 0
    enum_literals: tuple[str, ...] = ()
    record_fields: tuple[RttiField, ...] = ()

    @property
    def is_real(self) -> bool:
        return self.kind in {"sv_real", "vhdl_real"}

    @property
    def is_unpacked_array(self) -> bool:
        return self.kind == "sv_unpacked_array" or self.array_flavor == "unpacked"

    @property
    def is_vhdl_array(self) -> bool:
        return self.kind == "vhdl_array"

    @property
    def is_vhdl_record(self) -> bool:
        return self.kind == "vhdl_record"

    @property
    def is_vhdl_enum(self) -> bool:
        return self.kind == "vhdl_enum"

    @property
    def is_vhdl(self) -> bool:
        return self.kind.startswith("vhdl_")

    @property
    def is_supported_waveform(self) -> bool:
        return self.kind in {
            "sv_digital", "sv_unpacked_array", "sv_real",
            "vhdl_logic", "vhdl_bit", "vhdl_vector", "vhdl_bit_vector",
            "vhdl_integer", "vhdl_time", "vhdl_real", "vhdl_boolean",
            "vhdl_enum", "vhdl_record", "vhdl_array",
        }


@dataclass(frozen=True)
class DbgRange:
    index: int
    words: tuple[int, int, int, int, int, int]

    @property
    def count(self) -> int:
        return int(self.words[5])

    @property
    def left(self) -> int:
        return int(self.words[0])

    @property
    def step(self) -> int:
        raw = self.words[4]
        if raw == 1:
            return 1
        if raw == 0xFFFFFFFF:
            return -1
        raise DecodeError(f"unsupported DBG range direction 0x{raw:X} at range {self.index}")

    def indices(self) -> list[int]:
        return [self.left + self.step * i for i in range(self.count)]


@dataclass
class DbgObject:
    index: int
    name: str
    owner: int
    width: int
    type_code: int
    signed_code: int
    flags: int
    runtime_addr: int
    aux_addr: int
    alloc_flags: int
    object_index_field: int
    port_index: int
    type_info: RttiType | None = None
    dimensions: tuple[DbgRange, ...] = ()
    scopes: tuple[str, ...] = ()
    path: str = ""
    meta_words: tuple[int, ...] = ()
    alloc_words: tuple[int, ...] = ()
    sv_unpacked_rank: int = 0
    sv_packed_rank: int = 0
    sv_leaf_type: RttiType | None = None
    vhdl_array_rank: int = 0
    vhdl_leaf_type: RttiType | None = None

    @property
    def storage_kind(self) -> str:
        return self.type_info.kind if self.type_info else "unsupported"

    @property
    def is_real(self) -> bool:
        return bool(self.type_info and self.type_info.is_real)

    @property
    def is_unpacked_array(self) -> bool:
        return bool(self.type_info and self.type_info.is_unpacked_array)

    @property
    def is_vhdl(self) -> bool:
        return bool(self.type_info and self.type_info.is_vhdl)

    @property
    def bit_width(self) -> int:
        kind = self.storage_kind
        if kind in {"vhdl_integer"}:
            return 32
        if kind == "vhdl_time":
            return 64
        if kind in {"sv_real", "vhdl_real"}:
            return 64
        if kind == "vhdl_boolean":
            return 1
        if kind == "vhdl_enum":
            n = len(self.type_info.enum_literals) if self.type_info else 0
            return max(1, (max(1, n) - 1).bit_length())
        return max(1, self.width)

    @property
    def expected_payload_bytes(self) -> int:
        kind = self.storage_kind
        if kind in {"sv_real", "vhdl_real"}:
            return 8
        if kind in {"vhdl_logic", "vhdl_bit", "vhdl_boolean"}:
            return 1
        if kind in {"vhdl_vector", "vhdl_bit_vector"}:
            return max(1, self.width)
        if kind == "vhdl_integer":
            return 4
        if kind == "vhdl_time":
            return 8
        if kind == "vhdl_enum":
            return 1
        if kind == "vhdl_record":
            # Characterized DBG 006 records expose their padded runtime storage
            # byte count in the object width field (record_v: 5 data bytes -> 8).
            return max(1, self.width)
        if kind == "vhdl_array":
            shape = self.vhdl_array_leaf_shape
            if shape is None:
                raise DecodeError(f"VHDL array {self.path} has no leaf shape")
            leaves, _leaf_type, _leaf_width, leaf_payload, _dims = shape
            return len(leaves) * leaf_payload
        if kind == "sv_unpacked_array":
            shape = self.unpacked_leaf_shape
            if shape is None:
                raise DecodeError(f"unpacked array {self.path} has no leaf shape")
            leaves, _leaf_width, leaf_is_real, _packed = shape
            if leaf_is_real:
                return 8 * len(leaves)
            words = max(1, (self.bit_width + 31) // 32)
            return 8 * words
        if kind == "sv_digital":
            words = max(1, (self.bit_width + 31) // 32)
            return 8 * words
        raise DecodeError(
            f"unsupported RTTI waveform storage for {self.path}: "
            f"type={self.type_info.name if self.type_info else '?'} "
            f"tag=0x{self.type_info.tag:X}" if self.type_info else
            f"unsupported RTTI waveform storage for {self.path}: no type metadata"
        )

    @property
    def is_parameter(self) -> bool:
        return self.alloc_flags == 2 or self.flags == 1 or self.flags in {18, 19}

    @property
    def is_signed(self) -> bool:
        tname = self.type_info.name.lower() if self.type_info else ""
        return bool(self.signed_code != 0 or tname in {"signed", "integer"})

    @property
    def vcd_var_type(self) -> str:
        if self.is_real:
            return "real"
        if self.is_parameter:
            return "parameter"
        tname = (self.type_info.name.lower() if self.type_info else "")
        if tname in {"integer", "int", "natural", "positive"} and self.bit_width == 32:
            return "integer"
        if tname == "time" and self.bit_width == 64:
            return "time"
        if self.flags == 6:
            return "tri"
        if self.flags in {3, 14}:
            return "wire" if self.flags == 3 or self.is_vhdl else "reg"
        return "reg"

    @property
    def unpacked_leaf_shape(self) -> tuple[list[tuple[int, ...]], int, bool, tuple[DbgRange, ...]] | None:
        """Return (unpacked index tuples, leaf width, leaf-is-real, packed dimensions).

        DBG 006 stores all declared dimensions on the object, while the anonymous
        SystemVerilog RTTI type chain identifies which leading dimensions are
        unpacked.  Following that chain lets us flatten 1-D and multidimensional
        unpacked arrays without guessing from aggregate widths.
        """
        if not self.is_unpacked_array:
            return None
        rank = self.sv_unpacked_rank
        if rank <= 0 or len(self.dimensions) < rank:
            raise DecodeError(
                f"unpacked array {self.path} has inconsistent RTTI/DBG dimensions "
                f"(rank={rank}, descriptors={len(self.dimensions)})"
            )
        index_lists = [d.indices() for d in self.dimensions[:rank]]
        if any(not x for x in index_lists):
            raise DecodeError(f"unpacked array {self.path} has an empty index dimension")
        tuples: list[tuple[int, ...]] = [()]
        for indices in index_lists:
            tuples = [prefix + (idx,) for prefix in tuples for idx in indices]

        leaf = self.sv_leaf_type
        leaf_is_real = bool(leaf and leaf.kind == "sv_real")
        packed_dims = self.dimensions[rank:]
        if leaf_is_real:
            if packed_dims:
                raise DecodeError(
                    f"unpacked real array {self.path} unexpectedly has packed dimensions"
                )
            leaf_width = 64
        else:
            leaf_width = 1
            for d in packed_dims:
                leaf_width *= d.count
            if not packed_dims:
                if not tuples or self.width % len(tuples):
                    raise DecodeError(
                        f"unpacked array {self.path} aggregate width {self.width} is not divisible "
                        f"by {len(tuples)} leaf elements"
                    )
                leaf_width = self.width // len(tuples)
            expected_width = leaf_width * len(tuples)
            if expected_width != self.width:
                raise DecodeError(
                    f"unpacked array {self.path} DBG width {self.width} disagrees with "
                    f"RTTI/dimension-derived width {expected_width}"
                )
        return tuples, leaf_width, leaf_is_real, packed_dims

    @property
    def unpacked_element_shape(self) -> tuple[list[int], int] | None:
        """Backward-compatible 1-D view used only by older callers."""
        shape = self.unpacked_leaf_shape
        if shape is None:
            return None
        tuples, width, _real, _packed = shape
        if any(len(t) != 1 for t in tuples):
            raise DecodeError(f"{self.path} is multidimensional; use unpacked_leaf_shape")
        return [t[0] for t in tuples], width

    @property
    def vhdl_array_leaf_shape(self) -> tuple[list[tuple[int, ...]], RttiType, int, int, tuple[DbgRange, ...]] | None:
        """Return VHDL aggregate leaf layout from RTTI + DBG dimensions.

        The characterized VHDL array RTTI supplies the number of aggregate
        dimensions and the recursive element type.  DBG supplies concrete ranges.
        XSim stores std_logic/bit elements as one byte each in declaration order.
        """
        if not (self.type_info and self.type_info.kind == "vhdl_array"):
            return None
        rank = self.vhdl_array_rank
        if rank <= 0 or len(self.dimensions) < rank or self.vhdl_leaf_type is None:
            raise DecodeError(
                f"VHDL array {self.path} has inconsistent RTTI/DBG dimensions "
                f"(rank={rank}, descriptors={len(self.dimensions)})"
            )
        index_lists = [d.indices() for d in self.dimensions[:rank]]
        if any(not x for x in index_lists):
            raise DecodeError(f"VHDL array {self.path} has an empty index dimension")
        tuples: list[tuple[int, ...]] = [()]
        for indices in index_lists:
            tuples = [prefix + (idx,) for prefix in tuples for idx in indices]
        leaf = self.vhdl_leaf_type
        leaf_dims = self.dimensions[rank:]
        if leaf.kind in {"vhdl_vector", "vhdl_bit_vector"}:
            if not leaf_dims:
                raise DecodeError(f"VHDL vector-array leaf {self.path} has no concrete element range")
            leaf_width = 1
            for d in leaf_dims:
                leaf_width *= d.count
            leaf_payload = leaf_width
        elif leaf.kind in {"vhdl_logic", "vhdl_bit", "vhdl_boolean", "vhdl_enum"}:
            if leaf_dims:
                raise DecodeError(f"VHDL scalar-array leaf {self.path} unexpectedly has dimensions")
            leaf_width = max(1, (len(leaf.enum_literals)-1).bit_length()) if leaf.kind == "vhdl_enum" else 1
            leaf_payload = 1
        elif leaf.kind == "vhdl_integer":
            if leaf_dims:
                raise DecodeError(f"VHDL integer-array leaf {self.path} unexpectedly has dimensions")
            leaf_width, leaf_payload = 32, 4
        elif leaf.kind == "vhdl_time":
            if leaf_dims:
                raise DecodeError(f"VHDL time-array leaf {self.path} unexpectedly has dimensions")
            leaf_width, leaf_payload = 64, 8
        elif leaf.kind == "vhdl_real":
            if leaf_dims:
                raise DecodeError(f"VHDL real-array leaf {self.path} unexpectedly has dimensions")
            leaf_width, leaf_payload = 64, 8
        else:
            raise DecodeError(
                f"VHDL array {self.path} has uncharacterized leaf type "
                f"{leaf.name!r}/{leaf.kind}"
            )
        return tuples, leaf, leaf_width, leaf_payload, leaf_dims


@dataclass
class DebugDB:
    offset: int
    header_words: list[int]
    rtti_types: list[RttiType]
    ranges: list[DbgRange]
    structures: list[DbgStructure]
    objects: list[DbgObject]
    objects_by_addr: dict[int, list[DbgObject]]
    runtime_starts: list[int] = field(default_factory=list)
    runtime_ends: list[int] = field(default_factory=list)
    span_cache: dict[tuple[int, int], list[int]] = field(default_factory=dict)

    @property
    def top_modules(self) -> list[str]:
        return [s.name for s in self.structures if s.index != 0 and s.parent == 0]


@dataclass(frozen=True)
class WcfgSignal:
    path: str
    canonical: str
    style: str | None = None
    color: str | None = None
    radix: str | None = None


@dataclass
class WcfgInfo:
    path: Path
    signals: list[WcfgSignal]
    db_refs: list[str]


@dataclass(frozen=True)
class BatchJob:
    """One independently loadable conversion input in a GUI/CLI batch."""
    inputs: tuple[Path, ...]
    explicit_wcfg: Path | None
    source_hint: Path


@dataclass
class BatchExportResult:
    successes: list[tuple[Path, int, Path | None]]
    failures: list[tuple[Path, str]]
    report_path: Path
    fmt: str


@dataclass(slots=True)
class RawEvent:
    time: int
    payload: bytes
    chunk_serial: int
    slot: int


@dataclass(slots=True)
class RawFragment:
    root_index: int
    storage_id: int
    time: int
    payload: bytes
    chunk_serial: int
    slot: int


@dataclass
class Stream:
    root_index: int
    storage_id: int
    payload_bytes: int
    events: list[RawEvent] = field(default_factory=list)
    fragments: list[RawFragment] = field(default_factory=list)
    absolute_addr: int | None = None
    dbg_objects: list[DbgObject] = field(default_factory=list)
    selected_objects: list[DbgObject] = field(default_factory=list)

    def observed_digital_width(self) -> int:
        if self.payload_bytes % 8 != 0:
            return 1 << 30
        nwords = self.payload_bytes // 8
        maxbit = 0
        for e in self.events:
            if len(e.payload) != self.payload_bytes:
                continue
            for wi in range(nwords):
                aval, bval = struct.unpack_from("<II", e.payload, wi * 8)
                x = aval | bval
                if x:
                    maxbit = max(maxbit, wi * 32 + x.bit_length())
        return max(1, maxbit)

    @property
    def storage_kind(self) -> str:
        kinds = {o.storage_kind for o in self.dbg_objects}
        if len(kinds) != 1:
            raise DecodeError(
                f"runtime address 0x{int(self.absolute_addr or 0):X} maps to conflicting storage kinds: "
                + ", ".join(sorted(kinds))
            )
        return next(iter(kinds))

    @property
    def is_real(self) -> bool:
        kinds = {o.is_real for o in self.dbg_objects}
        if len(kinds) != 1:
            raise DecodeError(
                f"runtime address 0x{int(self.absolute_addr or 0):X} maps to conflicting real/digital DBG objects"
            )
        return next(iter(kinds))

    @property
    def width(self) -> int:
        if not self.dbg_objects:
            raise DecodeError("stream has no DBG object mapping")
        widths = {o.bit_width for o in self.dbg_objects}
        if len(widths) != 1:
            raise DecodeError(
                f"runtime address 0x{int(self.absolute_addr or 0):X} maps to conflicting DBG widths: {sorted(widths)}"
            )
        return next(iter(widths))


@dataclass
class RootInfo:
    index: int
    first_page: int
    pages: list[int] = field(default_factory=list)
    chunks: int = 0
    candidate_bases: list[int] = field(default_factory=list)
    bank_base: int | None = None


@dataclass
class WdbInfo:
    path: Path
    data: bytes
    final_section: int
    final_section_magic: str
    event_wrapper: int
    bank_size: int
    roots: list[RootInfo]
    fragments: list[RawFragment]
    streams: list[Stream]


# ---------------------------------------------------------------------------
# DBG parsing
# ---------------------------------------------------------------------------

def _cstring(pool: bytes, off: int) -> str:
    if off < 0 or off >= len(pool):
        return f"__bad_string_offset_0x{off:X}"
    z = pool.find(b"\0", off)
    if z < 0:
        z = len(pool)
    return pool[off:z].decode("utf-8", errors="replace")


def _structure_scope_chain(structures: list[DbgStructure], owner: int) -> tuple[str, ...]:
    names: list[str] = []
    seen: set[int] = set()
    while owner > 0:
        if owner in seen or owner >= len(structures):
            raise DecodeError(f"invalid/cyclic DBG structure parent chain at structure {owner}")
        seen.add(owner)
        s = structures[owner]
        names.append(s.name)
        owner = s.parent
    names.reverse()
    return tuple(names)


def parse_rtti_types(b: bytes, log: Logger) -> list[RttiType]:
    """Parse the embedded XSim TYPE FILE 001 table used by DBG object type indices."""
    pos = b.find(RTTI_TYPE_MAGIC)
    if pos < 0:
        raise DecodeError("embedded 'Xilinx ISim TYPE FILE 001' RTTI table was not found")
    hdr = (pos + len(RTTI_TYPE_MAGIC) + 3) & ~3
    if hdr + 12 > len(b):
        raise DecodeError("truncated XSim TYPE FILE 001 header")
    _checksum, count, body_size = struct.unpack_from("<3I", b, hdr)
    if count <= 0 or count > 1_000_000:
        raise DecodeError(f"implausible RTTI type count: {count}")
    off = hdr + 12
    types: list[RttiType] = []
    for idx in range(count):
        if off + 8 > len(b):
            raise DecodeError(f"truncated RTTI type record {idx}")
        size, tag = struct.unpack_from("<2I", b, off)
        if size < 9 or off + size > len(b):
            raise DecodeError(f"invalid RTTI type record {idx} size={size}")
        raw = b[off:off + size]
        z = raw.find(b"\0", 8)
        if z < 0:
            raise DecodeError(f"RTTI type record {idx} has no terminated name")
        name = raw[8:z].decode("utf-8", errors="replace")
        lname = name.lower()
        # XSim RTTI records carry a front-end/language code immediately after
        # named scalar types: 2 for VHDL and 5 for Verilog/SystemVerilog in all
        # characterized DBG 006 datasets. Anonymous SV composite types do not
        # carry this field in the same position, so exact type/tag recognition is
        # used for VHDL and the existing SV representation remains the default.
        lang_code = None
        if z + 5 <= len(raw):
            try:
                lang_code = struct.unpack_from("<I", raw, z + 1)[0]
            except struct.error:
                pass
        language = "vhdl" if lang_code == 2 else "sv" if lang_code == 5 else "unknown"
        kind = "sv_digital"
        array_flavor: str | None = None
        element_type_index: int | None = None
        array_rank = 0
        enum_literals: tuple[str, ...] = ()
        record_fields: tuple[RttiField, ...] = ()
        # Anonymous SystemVerilog array RTTI records are recursive.  The first
        # three u32 fields after the empty-name terminator are: version=1,
        # array descriptor (0x00A00002 unpacked / 0x00A00003 packed), and the
        # element RTTI index.  This is what distinguishes e.g. real[0:2] from a
        # 96-bit digital aggregate and reveals multidimensional unpacked rank.
        if tag == 0xA0000010 and name == "":
            tail = raw[z + 1:]
            if len(tail) >= 12:
                ver, descriptor, elem_idx = struct.unpack_from("<3I", tail, 0)
                if ver == 1 and descriptor in {0x00A00002, 0x00A00003}:
                    array_flavor = "unpacked" if descriptor == 0x00A00002 else "packed"
                    if elem_idx < count:
                        element_type_index = elem_idx

        # Named VHDL composite/enum RTTI is self-describing as well.  These
        # structures were characterized from XSim 2026.1 DBG 006 rather than
        # inferred from waveform values:
        #   enum   A0000003: lang, class, scalar-code, literal-count, strings...
        #   array  A0000010: lang, A00001, element-type-index, rank, ...
        #   record A0000011: lang, 000B0001, field-count, then per-field
        #                    name, type-index, range-count, (left,right,step)*N.
        tail = raw[z + 1:]
        if language == "vhdl" and tag == 0xA0000003 and len(tail) >= 16:
            _lang, _enum_class, _scalar_code, literal_count = struct.unpack_from("<4I", tail, 0)
            if 0 < literal_count <= 65536:
                cur = 16
                lits: list[str] = []
                ok = True
                for _ in range(literal_count):
                    zz = tail.find(b"\0", cur)
                    if zz < 0:
                        ok = False
                        break
                    lits.append(tail[cur:zz].decode("utf-8", errors="replace"))
                    cur = zz + 1
                if ok:
                    enum_literals = tuple(lits)

        if language == "vhdl" and tag == 0xA0000010 and len(tail) >= 16:
            _lang, descriptor, elem_idx, rank = struct.unpack_from("<4I", tail, 0)
            if descriptor == 0x00A00001 and elem_idx < count and 0 < rank <= 32:
                array_flavor = "vhdl_array"
                element_type_index = elem_idx
                array_rank = rank

        if language == "vhdl" and tag == 0xA0000011 and len(tail) >= 12:
            _lang, descriptor, field_count = struct.unpack_from("<3I", tail, 0)
            if descriptor == 0x000B0001 and 0 < field_count <= 4096:
                cur = 12
                fields: list[RttiField] = []
                ok = True
                for _ in range(field_count):
                    zz = tail.find(b"\0", cur)
                    if zz < 0:
                        ok = False
                        break
                    fname = tail[cur:zz].decode("utf-8", errors="replace")
                    cur = zz + 1
                    if cur + 8 > len(tail):
                        ok = False
                        break
                    ftype, rcount = struct.unpack_from("<2I", tail, cur)
                    cur += 8
                    if ftype >= count or rcount > 32 or cur + 12 * rcount > len(tail):
                        ok = False
                        break
                    franges: list[tuple[int, int, int]] = []
                    for _r in range(rcount):
                        left, right, step_raw = struct.unpack_from("<3I", tail, cur)
                        cur += 12
                        step = -1 if step_raw == 0xFFFFFFFF else int(step_raw)
                        if step not in {-1, 1}:
                            ok = False
                            break
                        franges.append((int(left), int(right), step))
                    if not ok:
                        break
                    fields.append(RttiField(fname, int(ftype), tuple(franges)))
                if ok and len(fields) == field_count:
                    record_fields = tuple(fields)

        uname = name.upper()
        # VHDL TIME uses a physical-type RTTI record whose field immediately
        # after the name is the unit count rather than the usual language code.
        # The uppercase name + physical-type tag is unambiguous in DBG 006.
        if tag == 0xA000000D and uname == "TIME":
            language = "vhdl"
            kind = "vhdl_time"
        elif language == "vhdl":
            if tag == 0xA0000006 and uname in {"REAL"}:
                kind = "vhdl_real"
            elif tag == 0xA0000003 and uname in {"STD_LOGIC", "STD_ULOGIC"}:
                kind = "vhdl_logic"
            elif tag == 0xA0000003 and uname == "BIT":
                kind = "vhdl_bit"
            elif tag == 0xA0000003 and uname == "BOOLEAN":
                kind = "vhdl_boolean"
            elif tag == 0xA0000010 and uname in {
                "STD_LOGIC_VECTOR", "STD_ULOGIC_VECTOR", "SIGNED", "UNSIGNED",
                "UNRESOLVED_SIGNED", "UNRESOLVED_UNSIGNED",
            }:
                kind = "vhdl_vector"
            elif tag == 0xA0000010 and uname == "BIT_VECTOR":
                kind = "vhdl_bit_vector"
            elif tag == 0xA0000005 and uname in {"INTEGER", "NATURAL", "POSITIVE"}:
                kind = "vhdl_integer"
            elif tag == 0xA0000003 and enum_literals:
                kind = "vhdl_enum"
            elif tag == 0xA0000011 and record_fields:
                kind = "vhdl_record"
            elif tag == 0xA0000010 and array_flavor == "vhdl_array" and element_type_index is not None:
                kind = "vhdl_array"
            else:
                # Files/access types and uncharacterized VHDL structures remain
                # fail-closed until their WDB storage has been characterized.
                kind = "unsupported"
        else:
            if tag == 0xA0000006 and lname in {"real", "realtime"}:
                kind = "sv_real"
            elif array_flavor == "unpacked":
                kind = "sv_unpacked_array"
            else:
                kind = "sv_digital"

        types.append(RttiType(idx, size, tag, name, raw, kind, language, array_flavor,
                              element_type_index, array_rank, enum_literals, record_fields))
        off += size

    # body_size is an implementation detail, but it must not point outside the file.
    if body_size and pos + body_size > len(b) + 64:
        raise DecodeError("RTTI TYPE FILE body size points outside the WDB")
    log.add("=== Embedded RTTI TYPE FILE 001 ===")
    log.add(f"RTTI offset           : 0x{pos:X}")
    log.add(f"type records          : {len(types)}")
    reals = [f"{t.index}:{t.name}/{t.kind}" for t in types if t.is_real]
    unpacked = [str(t.index) for t in types if t.is_unpacked_array]
    vhdl = [f"{t.index}:{t.name}/{t.kind}" for t in types if t.is_vhdl and t.is_supported_waveform]
    log.add("real types            : " + (", ".join(reals) if reals else "(none)"))
    log.add("unpacked-array types  : " + (", ".join(unpacked) if unpacked else "(none)"))
    log.add("supported VHDL types  : " + (", ".join(vhdl) if vhdl else "(none)"))
    return types


def parse_debug_db(b: bytes, log: Logger) -> DebugDB:
    off = b.find(DBG_MAGIC)
    if off < 0:
        pos = b.find(b"Xilinx ISim DBG ")
        if pos >= 0:
            z = b.find(b"\0", pos, min(len(b), pos + 64))
            found = b[pos:z if z >= 0 else pos + 32].decode("ascii", errors="replace")
            raise DecodeError(f"unsupported embedded DBG version: {found!r}; this decoder requires DBG 006")
        raise DecodeError("embedded 'Xilinx ISim DBG 006' database was not found")

    rtti_types = parse_rtti_types(b, log)

    base = off + len(DBG_MAGIC)
    if base + 35 * 4 > len(b):
        raise DecodeError("truncated DBG 006 header")
    w = [u32(b, base + 4 * i) for i in range(35)]

    struct_table = off + w[2]
    alloc_table = off + w[4]
    obj_table = off + w[5]
    range_table = off + w[6]
    struct_pool_start = off + w[7]
    signal_pool_start = off + w[12]
    signal_pool_end = off + w[13]
    struct_count = w[20]
    object_count = w[22]
    range_count = w[24]

    if struct_count <= 0 or struct_count > 1_000_000:
        raise DecodeError(f"implausible DBG structure count: {struct_count}")
    if object_count <= 0 or object_count > 10_000_000:
        raise DecodeError(f"implausible DBG object count: {object_count}")
    if range_count > 10_000_000:
        raise DecodeError(f"implausible DBG range count: {range_count}")

    checks = [
        (struct_table, struct_count * STRUCT_REC_SIZE, "structure table"),
        (obj_table, object_count * OBJ_REC_SIZE, "object table"),
        (alloc_table, object_count * ALLOC_REC_SIZE, "allocation table"),
        (range_table, range_count * 24, "range table"),
    ]
    for table_start, size, label in checks:
        if table_start < off or table_start + size > len(b):
            raise DecodeError(f"DBG {label} is outside the WDB")
    if not (off <= struct_pool_start <= signal_pool_start <= signal_pool_end <= len(b)):
        raise DecodeError("DBG string-pool bounds are invalid")

    struct_pool = b[struct_pool_start:signal_pool_start]
    signal_pool = b[signal_pool_start:signal_pool_end]

    structures: list[DbgStructure] = []
    for i in range(struct_count):
        vals = struct.unpack_from("<9I", b, struct_table + i * STRUCT_REC_SIZE)
        structures.append(DbgStructure(i, _cstring(struct_pool, vals[0]), vals[1]))

    ranges: list[DbgRange] = []
    for i in range(range_count):
        vals = struct.unpack_from("<6I", b, range_table + i * 24)
        r = DbgRange(i, vals)
        if r.count <= 0 or r.count > 100_000_000:
            raise DecodeError(f"implausible DBG range size {r.count} at range {i}")
        ranges.append(r)

    objects: list[DbgObject] = []
    by_addr: dict[int, list[DbgObject]] = defaultdict(list)
    for i in range(object_count):
        m = struct.unpack_from("<11I", b, obj_table + i * OBJ_REC_SIZE)
        a = struct.unpack_from("<14I", b, alloc_table + i * ALLOC_REC_SIZE)
        owner = a[4]
        scopes = _structure_scope_chain(structures, owner)
        name = _cstring(signal_pool, m[0])
        path = "/" + "/".join((*scopes, name))
        runtime_addr = a[0]
        type_index = m[5]
        type_info: RttiType | None
        if type_index < len(rtti_types):
            type_info = rtti_types[type_index]
        else:
            # DBG 006 also contains compiler/package bookkeeping entries that are
            # not waveform objects (common in VHDL ieee.numeric_std). They may use
            # 0xFFFFFFFF sentinels for type/range fields. Only a runtime allocation
            # is allowed to participate in WDB event mapping, so invalid metadata is
            # tolerated solely for address-zero bookkeeping entries.
            type_info = None

        dim_count = m[6]
        dim_start = m[7]
        dims: tuple[DbgRange, ...] = ()
        if dim_count and dim_count != 0xFFFFFFFF:
            valid_dim_span = (
                dim_start != 0xFFFFFFFF and
                dim_start <= len(ranges) and
                dim_count <= len(ranges) - dim_start
            )
            if not valid_dim_span:
                # Non-waveform compiler metadata may use sentinel range indices.
                # If such an allocation is ever targeted by a WDB event, mapping
                # validation below will reject it instead of assigning a name.
                pass
            else:
                dims = tuple(ranges[dim_start:dim_start + dim_count])
                # Range descriptors on VHDL package/function bookkeeping entries
                # are not always HDL bit dimensions. Defer strict dimension-width
                # validation until an allocation is actually mapped to WDB events.
        sv_unpacked_rank = 0
        sv_packed_rank = 0
        sv_leaf_type = type_info
        seen_type_indices: set[int] = set()
        while sv_leaf_type is not None and sv_leaf_type.array_flavor in {"unpacked", "packed"}:
            if sv_leaf_type.index in seen_type_indices:
                raise DecodeError(f"cyclic RTTI array type chain at type {sv_leaf_type.index}")
            seen_type_indices.add(sv_leaf_type.index)
            if sv_leaf_type.array_flavor == "unpacked":
                sv_unpacked_rank += 1
            else:
                sv_packed_rank += 1
            nxt = sv_leaf_type.element_type_index
            if nxt is None or nxt >= len(rtti_types):
                raise DecodeError(f"RTTI array type {sv_leaf_type.index} has invalid element type")
            sv_leaf_type = rtti_types[nxt]

        vhdl_array_rank = 0
        vhdl_leaf_type = type_info
        seen_vhdl_indices: set[int] = set()
        while vhdl_leaf_type is not None and vhdl_leaf_type.kind == "vhdl_array":
            if vhdl_leaf_type.index in seen_vhdl_indices:
                raise DecodeError(f"cyclic VHDL RTTI array type chain at type {vhdl_leaf_type.index}")
            seen_vhdl_indices.add(vhdl_leaf_type.index)
            if vhdl_leaf_type.array_rank <= 0:
                raise DecodeError(f"VHDL RTTI array type {vhdl_leaf_type.index} has invalid rank")
            vhdl_array_rank += vhdl_leaf_type.array_rank
            nxt = vhdl_leaf_type.element_type_index
            if nxt is None or nxt >= len(rtti_types):
                raise DecodeError(f"VHDL RTTI array type {vhdl_leaf_type.index} has invalid element type")
            vhdl_leaf_type = rtti_types[nxt]

        obj = DbgObject(
            index=i,
            name=name,
            owner=owner,
            width=m[4],
            type_code=type_index,
            signed_code=m[1],
            flags=m[8],
            runtime_addr=runtime_addr,
            aux_addr=a[2],
            alloc_flags=a[7],
            object_index_field=a[8],
            port_index=a[10],
            type_info=type_info,
            dimensions=dims,
            scopes=scopes,
            path=path,
            meta_words=m,
            alloc_words=a,
            sv_unpacked_rank=sv_unpacked_rank,
            sv_packed_rank=sv_packed_rank,
            sv_leaf_type=sv_leaf_type,
            vhdl_array_rank=vhdl_array_rank,
            vhdl_leaf_type=vhdl_leaf_type,
        )
        # Validate the supported unpacked-array storage shape.
        if obj.is_unpacked_array and obj.runtime_addr and obj.width > 0:
            _ = obj.unpacked_leaf_shape
        objects.append(obj)
        # Some VHDL package/function bookkeeping entries have nonzero allocation
        # addresses and perfectly valid RTTI type IDs, but no concrete object
        # ranges (e.g. numeric_std function locals).  Treat an object as a WDB
        # mapping candidate only if its exact supported storage shape can be
        # derived.  If an event ever targets an excluded object, root resolution
        # will fail closed rather than guessing.
        candidate = (
            obj.runtime_addr and obj.type_info is not None and obj.type_info.is_supported_waveform
            and 0 < obj.width < 0xFFFFFFFF
        )
        if candidate:
            try:
                _ = obj.expected_payload_bytes
            except DecodeError:
                candidate = False
        if candidate:
            by_addr[obj.runtime_addr].append(obj)

    log.add("\n=== Embedded DBG 006 ===")
    log.add(f"DBG offset            : 0x{off:X}")
    log.add(f"structures            : {len(structures)}")
    log.add(f"objects               : {len(objects)}")
    log.add(f"runtime addresses     : {len(by_addr)}")
    log.add(f"range descriptors     : {len(ranges)}")
    tops = [s.name for s in structures if s.index != 0 and s.parent == 0]
    log.add("top modules           : " + (", ".join(tops) if tops else "(none identified)"))
    log.add(f"structure table       : 0x{struct_table:X}")
    log.add(f"object table          : 0x{obj_table:X}")
    log.add(f"allocation table      : 0x{alloc_table:X}")
    log.add(f"range table           : 0x{range_table:X}")

    # Keep the sorted allocation bounds for diagnostics. VHDL DBG files can
    # contain overlapping compiler/package-local bookkeeping allocations, so
    # overlap is not rejected globally. Actual WDB fragment mapping requires a
    # unique containing waveform allocation and will reject ambiguity there.
    starts: list[int] = []
    ends: list[int] = []
    for addr in sorted(by_addr):
        objs = by_addr[addr]
        sizes = {o.expected_payload_bytes for o in objs}
        realness = {o.is_real for o in objs}
        widths = {o.bit_width for o in objs}
        if len(sizes) != 1 or len(realness) != 1 or len(widths) != 1:
            # Conflicting same-address aliases cannot be waveform-mapped safely.
            # Preserve the DBG objects, but omit this address from the span index.
            continue
        size = next(iter(sizes))
        starts.append(addr)
        ends.append(addr + size)

    return DebugDB(off, w, rtti_types, ranges, structures, objects, dict(by_addr), starts, ends, {})


# ---------------------------------------------------------------------------
# WCFG parsing / path normalization
# ---------------------------------------------------------------------------

def canonical_path(path: str) -> str:
    # WCFG escaped identifiers use a leading backslash and trailing space,
    # e.g. /top/\\g_nodes[0].u_leaf /clk. DBG stores the same component as
    # g_nodes[0].u_leaf. Canonicalization makes them comparable.
    comps: list[str] = []
    for raw in path.replace("\\\\", "\\").split("/"):
        if not raw:
            continue
        c = raw.strip()
        if c.startswith("\\"):
            c = c[1:].rstrip()
        if c:
            comps.append(c)
    return "/" + "/".join(comps)


def normalize_wcfg_radix(value: str | None) -> str | None:
    """Translate Vivado WCFG Radix properties to table-export formats."""
    if not value:
        return None
    v = value.strip().upper().replace("_", "")
    mapping = {
        "HEXRADIX": "hex",
        "UNSIGNEDDECRADIX": "int",
        "SIGNEDDECRADIX": "signed",
        "BINARYRADIX": "bin",
        "BINRADIX": "bin",
        "OCTALRADIX": "oct",
        "OCTRADIX": "oct",
    }
    return mapping.get(v)


def parse_wcfg(path: Path) -> WcfgInfo:
    try:
        root = ET.parse(path).getroot()
    except (ET.ParseError, OSError) as exc:
        raise DecodeError(f"cannot parse WCFG {path}: {exc}") from exc

    signals: list[WcfgSignal] = []
    refs: list[str] = []
    seen: set[str] = set()
    for elem in root.iter():
        tag = elem.tag.rsplit("}", 1)[-1]
        if tag == "db_ref":
            p = elem.attrib.get("path")
            if p:
                refs.append(p)
        elif tag == "wvobject":
            typ = (elem.attrib.get("type") or "").lower()
            p = elem.attrib.get("fp_name")
            if not p or typ in {"group", "divider"}:
                continue
            props: dict[str, str] = {}
            for ch in elem:
                ctag = ch.tag.rsplit("}", 1)[-1]
                if ctag == "obj_property":
                    props[ch.attrib.get("name", "")] = (ch.text or "").strip()
            cp = canonical_path(p)
            if cp in seen:
                continue
            seen.add(cp)
            signals.append(WcfgSignal(
                path=p,
                canonical=cp,
                style=props.get("WaveformStyle"),
                color=props.get("CustomSignalColor") or props.get("Color"),
                radix=normalize_wcfg_radix(props.get("Radix")),
            ))
    return WcfgInfo(path.resolve(), signals, refs)


def resolve_wdb_from_wcfg(info: WcfgInfo) -> Path:
    # Same-stem beside WCFG is the most portable answer when a WCFG has stale
    # absolute project paths.
    same = info.path.with_suffix(".wdb")
    if same.exists():
        return same.resolve()

    for ref in info.db_refs:
        # Native Path handles absolute Windows paths when run on Windows.
        p = Path(ref)
        if not p.is_absolute():
            p = info.path.parent / p
        if p.exists():
            return p.resolve()
        # Also try the basename beside the WCFG.
        alt = info.path.parent / Path(ref.replace("\\", "/")).name
        if alt.exists():
            return alt.resolve()
    raise DecodeError("WCFG does not resolve to an existing WDB; pass the WDB explicitly")


def build_batch_jobs(source_paths: list[Path], wcfg_paths: list[Path] | None = None) -> list[BatchJob]:
    """Build deterministic WDB/WCFG or VCD jobs from multi-select input.

    WCFG files are paired to WDBs by resolved db_ref first, then by same stem.
    A single WDB + single WCFG pair is accepted even when the WCFG contains a
    stale absolute path, matching the normal explicit-pair workflow. Ambiguous
    pairings fail closed rather than silently applying the wrong WCFG.
    """
    wcfg_paths = list(wcfg_paths or [])
    src = [Path(p).resolve() for p in source_paths]
    extra_cfg = [Path(p).resolve() for p in wcfg_paths]
    if not src:
        raise ConversionError("no waveform files selected")
    for p in src + extra_cfg:
        if not p.is_file():
            raise ConversionError(f"input file not found: {p}")

    vcds = [p for p in src if p.suffix.lower() == ".vcd"]
    wdbs = [p for p in src if p.suffix.lower() == ".wdb"]
    src_cfgs = [p for p in src if p.suffix.lower() == ".wcfg"]
    bad = [p for p in src if p.suffix.lower() not in {".vcd", ".wdb", ".wcfg"}]
    if bad:
        raise ConversionError(f"unsupported waveform input: {bad[0]}")

    if vcds:
        if wdbs or src_cfgs or extra_cfg:
            raise ConversionError("VCD files cannot be mixed with WDB/WCFG files in one batch")
        return [BatchJob((p,), None, p) for p in vcds]

    cfgs: list[Path] = []
    seen_cfg: set[Path] = set()
    for p in src_cfgs + extra_cfg:
        if p not in seen_cfg:
            seen_cfg.add(p)
            cfgs.append(p)

    # WCFG-only batch: each config resolves its own WDB independently.
    if not wdbs:
        if not cfgs:
            raise ConversionError("select at least one WDB, WCFG or VCD")
        jobs: list[BatchJob] = []
        for cfg in cfgs:
            info = parse_wcfg(cfg)
            source = resolve_wdb_from_wcfg(info)
            jobs.append(BatchJob((cfg,), None, source))
        return jobs

    # Cache WCFG metadata once. A stale db_ref is allowed only if another
    # unambiguous pairing rule (same stem or the lone explicit pair) exists.
    cfg_info: dict[Path, tuple[WcfgInfo, Path | None]] = {}
    for cfg in cfgs:
        info = parse_wcfg(cfg)
        try:
            resolved = resolve_wdb_from_wcfg(info)
        except DecodeError:
            resolved = None
        cfg_info[cfg] = (info, resolved)

    used_cfg: set[Path] = set()
    jobs = []
    for wdb in wdbs:
        ranked: list[tuple[int, Path]] = []
        for cfg, (_info, resolved) in cfg_info.items():
            if resolved is not None and resolved == wdb:
                ranked.append((0, cfg))
            elif cfg.stem.casefold() == wdb.stem.casefold():
                ranked.append((1, cfg))
            elif len(wdbs) == 1 and len(cfgs) == 1:
                ranked.append((2, cfg))
        chosen: Path | None = None
        if ranked:
            best = min(rank for rank, _cfg in ranked)
            choices = [cfg for rank, cfg in ranked if rank == best]
            if len(choices) != 1:
                names = ", ".join(x.name for x in choices)
                raise ConversionError(f"ambiguous WCFG pairing for {wdb.name}: {names}")
            chosen = choices[0]
            if chosen in used_cfg:
                raise ConversionError(f"WCFG would be paired with more than one WDB: {chosen.name}")
            used_cfg.add(chosen)
        jobs.append(BatchJob((wdb,), chosen, wdb))

    unused = [cfg for cfg in cfgs if cfg not in used_cfg]
    if unused:
        names = ", ".join(x.name for x in unused)
        raise ConversionError(f"WCFG file(s) could not be paired to a selected WDB: {names}")
    return jobs


def unique_batch_output(outdir: Path, source: Path, fmt: str, used: set[Path]) -> Path:
    ext = {"vcd": ".vcd", "csv": ".csv", "json": ".json", "excel": ".xlsx"}[fmt]
    base = source.stem
    candidate = (outdir / f"{base}{ext}").resolve()
    serial = 2
    while candidate in used:
        candidate = (outdir / f"{base}_{serial}{ext}").resolve()
        serial += 1
    used.add(candidate)
    return candidate


# ---------------------------------------------------------------------------
# WDB event parsing
# ---------------------------------------------------------------------------

def _read_section_magic(b: bytes, off: int) -> str:
    if off < 0 or off >= len(b):
        return "<outside-file>"
    z = b.find(b"\0", off, min(len(b), off + 64))
    if z < 0:
        z = min(len(b), off + 32)
    return b[off:z].decode("ascii", errors="replace")


def parse_bank_size(b: bytes, log: Logger) -> tuple[int, int]:
    event_wrapper = b.find(EVENT_MAGIC)
    if event_wrapper < 0:
        raise DecodeError("WDB.Event wrapper was not found")
    if event_wrapper + 0x30 > len(b):
        raise DecodeError("truncated WDB.Event wrapper")
    meta_ptr = u64(b, event_wrapper + 0x20)
    meta_size = u64(b, event_wrapper + 0x28)
    bank = DEFAULT_BANK_SIZE
    if meta_ptr + meta_size <= len(b) and meta_size >= 20:
        candidate = u32(b, meta_ptr + 16)
        if is_power_of_two(candidate) and 0x100 <= candidate <= 0x100000:
            bank = candidate
    log.add(f"WDB.Event wrapper     : 0x{event_wrapper:X}")
    log.add(f"runtime bank size     : 0x{bank:X}")
    return event_wrapper, bank


def _page_valid(b: bytes, off: int, require_nonempty: bool = True) -> bool:
    if off < 0 or off + PAGE_SIZE > len(b):
        return False
    cnt = u32(b, off + COUNT_OFF)
    if cnt > MAX_CHUNKS_PER_PAGE or (require_nonempty and cnt == 0):
        return False
    for i in range(cnt):
        ptr = u64(b, off + PTR_ARRAY_OFF + i * 8)
        size = u32(b, off + SIZE_ARRAY_OFF + i * 4)
        if ptr <= 0 or size <= 0 or ptr + size > len(b):
            return False
    return True


def discover_roots(b: bytes, final_section: int) -> list[RootInfo]:
    start = final_section + ROOT0_REL
    roots: list[RootInfo] = []
    earliest_payload = len(b)
    for idx in range(1024):
        off = start + idx * PAGE_SIZE
        if off >= earliest_payload or not _page_valid(b, off, require_nonempty=True):
            break
        cnt = u32(b, off + COUNT_OFF)
        for i in range(cnt):
            earliest_payload = min(earliest_payload, u64(b, off + PTR_ARRAY_OFF + i * 8))
        roots.append(RootInfo(idx, off))
    if not roots:
        raise DecodeError(
            f"no event-root pages recognized at final WDB section +0x30 (0x{start:X})"
        )
    return roots


def iter_page_chain(b: bytes, first: int) -> Iterator[int]:
    seen: set[int] = set()
    page = first
    while page:
        if page in seen:
            raise DecodeError(f"cycle in event page chain at 0x{page:X}")
        seen.add(page)
        if not _page_valid(b, page, require_nonempty=False):
            raise DecodeError(f"invalid event index page at 0x{page:X}")
        yield page
        nxt = u64(b, page)
        if nxt and (nxt < 0 or nxt + PAGE_SIZE > len(b)):
            raise DecodeError(f"continuation page pointer 0x{nxt:X} is outside WDB")
        page = nxt


def decode_chunk(payload: bytes, chunk_serial: int, root_index: int,
                 fragments: list[RawFragment]) -> None:
    if len(payload) < CHUNK_HEADER_SIZE:
        raise DecodeError(f"decompressed chunk {chunk_serial} is only {len(payload)} bytes")
    record_count = u32(payload, 16)
    off = CHUNK_HEADER_SIZE
    for slot in range(record_count):
        if off + EVENT_FIXED_SIZE > len(payload):
            raise DecodeError(
                f"chunk {chunk_serial}: record {slot}/{record_count} header exceeds decompressed data"
            )
        tlo, thi, sid, pbytes = struct.unpack_from("<4I", payload, off)
        off += EVENT_FIXED_SIZE
        if pbytes == 0 or pbytes > 1_048_576:
            raise DecodeError(
                f"chunk {chunk_serial}: implausible payload_bytes={pbytes} for storage 0x{sid:X}"
            )
        # Do not impose an 8-byte alignment here. SystemVerilog aval/bval
        # records are word-aligned, but VHDL STD_LOGIC(_VECTOR) records use
        # one byte per element and can start at arbitrary byte offsets. The
        # exact storage representation is validated only after DBG/RTTI mapping.
        if off + pbytes > len(payload):
            raise DecodeError(
                f"chunk {chunk_serial}: record payload exceeds decompressed data (need {pbytes} bytes)"
            )
        value = payload[off:off + pbytes]
        off += pbytes
        fragments.append(
            RawFragment(root_index, sid, tlo | (thi << 32), value, chunk_serial, slot)
        )
    # XSim pads decompressed members to a fixed working-buffer size. Non-zero
    # trailing bytes would indicate that our record parser lost synchronization.
    if any(payload[off:]):
        raise DecodeError(
            f"chunk {chunk_serial}: {len(payload)-off} non-zero trailing bytes after "
            f"{record_count} records"
        )


def parse_wdb(path: Path, log: Logger) -> WdbInfo:
    b = path.read_bytes()
    if not b.startswith(WDB_MAGIC):
        raise DecodeError(f"unsupported WDB signature: {b[:32]!r}")
    if len(b) < 0x60:
        raise DecodeError("truncated WDB header")

    # WDB 01 stores three top-level section pointers at 0x48/0x50/0x58.
    # Their *order by type* is not fixed: characterized XSim outputs use both
    # RTTI->DBG->Event and Event->RTTI->DBG. What is stable is that the three
    # pointers are ordered by file position and identify exactly these wrappers.
    section_ptrs = [u64(b, off) for off in (0x48, 0x50, 0x58)]
    if any(off <= 0 or off >= len(b) for off in section_ptrs):
        raise DecodeError("one or more WDB section pointers are outside the file")
    if section_ptrs != sorted(section_ptrs) or len(set(section_ptrs)) != 3:
        raise DecodeError(
            "WDB section pointers are not three unique ascending offsets; refusing unknown container layout"
        )
    section_magics = [_read_section_magic(b, off) for off in section_ptrs]
    expected_sections = {"Xilinx RTTI", "Xilinx DBG", "WDB.Event"}
    if set(section_magics) != expected_sections or len(section_magics) != len(set(section_magics)):
        raise DecodeError(
            "unsupported WDB 01 section set: " + ", ".join(repr(x) for x in section_magics)
            + "; expected exactly 'Xilinx RTTI', 'Xilinx DBG', and 'WDB.Event'"
        )
    final_section = section_ptrs[-1]
    final_magic = section_magics[-1]
    event_header_ptr = section_ptrs[section_magics.index("WDB.Event")]
    event_wrapper, bank_size = parse_bank_size(b, log)
    if event_wrapper != event_header_ptr:
        raise DecodeError(
            f"WDB.Event pointer mismatch: header has 0x{event_header_ptr:X}, "
            f"wrapper scan found 0x{event_wrapper:X}; refusing to guess container layout"
        )
    roots = discover_roots(b, final_section)

    log.add("\n=== WDB container ===")
    log.add(f"file                  : {path.resolve()}")
    log.add(f"size                  : {len(b)} bytes (0x{len(b):X})")
    log.add(f"header section[0x48]  : 0x{u64(b,0x48):X} ({_read_section_magic(b,u64(b,0x48))})")
    log.add(f"header section[0x50]  : 0x{u64(b,0x50):X} ({_read_section_magic(b,u64(b,0x50))})")
    log.add(f"header section[0x58]  : 0x{final_section:X} ({final_magic})")
    log.add(f"event root table      : 0x{final_section + ROOT0_REL:X}")

    fragments: list[RawFragment] = []
    chunk_serial = 0
    log.add("\n=== Event roots / zlib chunks ===")
    for root in roots:
        before = len(fragments)
        for page in iter_page_chain(b, root.first_page):
            root.pages.append(page)
            cnt = u32(b, page + COUNT_OFF)
            for i in range(cnt):
                ptr = u64(b, page + PTR_ARRAY_OFF + i * 8)
                size = u32(b, page + SIZE_ARRAY_OFF + i * 4)
                try:
                    d = zlib.decompress(b[ptr:ptr + size])
                except zlib.error as exc:
                    raise DecodeError(
                        f"zlib failure root={root.index} page=0x{page:X} slot={i} "
                        f"ptr=0x{ptr:X} size={size}: {exc}"
                    ) from exc
                decode_chunk(d, chunk_serial, root.index, fragments)
                chunk_serial += 1
                root.chunks += 1
        rfrags = fragments[before:]
        shapes: dict[int, set[int]] = defaultdict(set)
        for fr in rfrags:
            shapes[fr.storage_id].add(len(fr.payload))
        shape_text = ", ".join(
            f"0x{sid:X}/" + "|".join(f"{n}B" for n in sorted(sizes))
            for sid, sizes in sorted(shapes.items())
        )
        log.add(
            f"root {root.index:3d}: first=0x{root.first_page:X} pages={len(root.pages):3d} "
            f"chunks={root.chunks:5d} records={len(rfrags):7d} starts={len(shapes):3d}  "
            + shape_text
        )

    if not fragments:
        raise DecodeError("WDB event table contained no waveform records")
    return WdbInfo(path.resolve(), b, final_section, final_magic, event_wrapper, bank_size,
                   roots, fragments, [])


# ---------------------------------------------------------------------------
# Exact event-root -> DBG runtime bank mapping
# ---------------------------------------------------------------------------

def digital_payload_words(payload: bytes) -> tuple[list[int], list[int]]:
    if len(payload) % 8:
        raise DecodeError(f"digital payload length {len(payload)} is not divisible by 8")
    avals: list[int] = []
    bvals: list[int] = []
    for off in range(0, len(payload), 8):
        a, b = struct.unpack_from("<II", payload, off)
        avals.append(a)
        bvals.append(b)
    return avals, bvals


def _alias_group_metadata(objs: list[DbgObject], addr: int) -> tuple[int, str]:
    if not objs:
        raise DecodeError(f"runtime address 0x{addr:X} has no DBG objects")
    try:
        sizes = {o.expected_payload_bytes for o in objs}
    except DecodeError:
        raise
    kinds = {o.storage_kind for o in objs}
    widths = {o.bit_width for o in objs}
    if len(sizes) != 1 or len(kinds) != 1 or len(widths) != 1:
        raise DecodeError(
            f"runtime address 0x{addr:X} has conflicting alias metadata: "
            + ", ".join(
                f"{o.path}[bits={o.bit_width},raw_width={o.width},type={o.type_code},"
                f"kind={o.storage_kind}]" for o in objs
            )
        )
    return next(iter(sizes)), next(iter(kinds))


def _span_object_bases(dbg: DebugDB, abs_start: int, payload_bytes: int) -> list[int]:
    """Return waveform-capable DBG allocations containing a WDB fragment.

    VHDL DBG metadata can contain overlapping package/compiler allocations, so
    we do not assume a globally non-overlapping address space. Results are
    cached by fragment span; WDBs typically have millions of records but only
    a small number of distinct (storage offset, payload size) shapes.
    """
    if payload_bytes <= 0:
        return []
    key = (abs_start, payload_bytes)
    cached = dbg.span_cache.get(key)
    if cached is not None:
        return cached
    end = abs_start + payload_bytes
    matches: list[int] = []
    for base, objs in dbg.objects_by_addr.items():
        try:
            size, _ = _alias_group_metadata(objs, base)
        except DecodeError:
            continue
        if base <= abs_start and end <= base + size:
            matches.append(base)
    matches.sort()
    dbg.span_cache[key] = matches
    return matches


def _root_fragments(wdb: WdbInfo, root_index: int) -> list[RawFragment]:
    return [f for f in wdb.fragments if f.root_index == root_index]


def root_candidate_bases(root: RootInfo, wdb: WdbInfo, dbg: DebugDB, bank_size: int) -> list[int]:
    rfrags = _root_fragments(wdb, root.index)
    shapes = sorted({(f.storage_id, len(f.payload)) for f in rfrags})
    bases = sorted({addr & ~(bank_size - 1) for addr in dbg.objects_by_addr if addr})
    out: list[int] = []
    for bank_base in bases:
        ok = True
        for sid, pbytes in shapes:
            starts = _span_object_bases(dbg, bank_base + sid, pbytes)
            if len(starts) != 1:
                ok = False
                break
        if ok:
            out.append(bank_base)
    return out


def enumerate_assignments(candidate_lists: list[list[int]], limit: int = 256) -> list[tuple[int, ...]]:
    order = sorted(range(len(candidate_lists)), key=lambda i: len(candidate_lists[i]))
    assignment: list[int | None] = [None] * len(candidate_lists)
    used: set[int] = set()
    solutions: list[tuple[int, ...]] = []

    def rec(k: int) -> None:
        if len(solutions) >= limit:
            return
        if k == len(order):
            solutions.append(tuple(int(x) for x in assignment if x is not None))
            return
        ri = order[k]
        for base in candidate_lists[ri]:
            if base in used:
                continue
            used.add(base)
            assignment[ri] = base
            rec(k + 1)
            assignment[ri] = None
            used.remove(base)

    rec(0)
    return solutions


def _selected_dbg_addresses(dbg: DebugDB, wcfg: WcfgInfo | None) -> set[int]:
    if not wcfg:
        return set()
    wanted = {s.canonical for s in wcfg.signals}
    out: set[int] = set()
    for o in dbg.objects:
        if not o.runtime_addr:
            continue
        cp = canonical_path(o.path)
        if cp in wanted:
            out.add(o.runtime_addr)
            continue
        if o.is_unpacked_array:
            shape = o.unpacked_leaf_shape
            if shape:
                index_tuples, _width, _real, _packed = shape
                if any(
                    canonical_path(o.path + "".join(f"[{idx}]" for idx in tup)) in wanted
                    for tup in index_tuples
                ):
                    out.add(o.runtime_addr)
    return out


def _solution_present_addresses(wdb: WdbInfo, dbg: DebugDB, sol: tuple[int, ...]) -> set[int]:
    present: set[int] = set()
    for root in wdb.roots:
        bank = sol[root.index]
        for fr in _root_fragments(wdb, root.index):
            starts = _span_object_bases(dbg, bank + fr.storage_id, len(fr.payload))
            if len(starts) == 1:
                present.add(starts[0])
    return present


def _header_root_bank_hints(wdb: WdbInfo, log: Logger) -> tuple[int, ...] | None:
    """Recover the supported WDB.Event root-to-runtime-bank table when present.

    When WDB.Event is the first top-level section, XSim WDB 01 stores sparse
    64-bit root-page pointers beginning at file offset 0xD0. Slot N corresponds
    to runtime bank (N+1)*bank_size. Older/alternate WDB 01 layouts place RTTI
    first and do not have this table in the file header; those files use the
    independent DBG-candidate resolver below.
    """
    if wdb.event_wrapper <= 0xD0:
        return None
    root_by_page = {r.first_page: r.index for r in wdb.roots}
    hits: dict[int, list[int]] = defaultdict(list)
    # The hint table, when present, is before the WDB.Event wrapper itself.
    # Scan only aligned qwords in that header region.
    stop = min(wdb.event_wrapper, len(wdb.data) - 7)
    for off in range(0xD0, stop, 8):
        value = u64(wdb.data, off)
        ri = root_by_page.get(value)
        if ri is None:
            continue
        slot = (off - 0xD0) // 8
        hits[ri].append((slot + 1) * wdb.bank_size)

    if not hits:
        return None
    if set(hits) != set(range(len(wdb.roots))) or any(len(v) != 1 for v in hits.values()):
        log.add("header bank hints      : partial/ambiguous; using DBG resolver")
        return None
    ordered = tuple(hits[i][0] for i in range(len(wdb.roots)))
    if len(set(ordered)) != len(ordered):
        raise DecodeError("WDB header root-bank hints assign more than one event root to the same runtime bank")
    log.add("header bank hints      : " + ", ".join(f"r{i}=0x{b:X}" for i, b in enumerate(ordered)))
    return ordered


def _validate_root_bank(root: RootInfo, bank_base: int, rfrags: list[RawFragment], dbg: DebugDB) -> None:
    # Validate only unique fragment shapes; millions of repeated event records
    # should not cause millions of allocation searches.
    shapes = {(f.storage_id, len(f.payload)) for f in rfrags}
    for sid, pbytes in shapes:
        starts = _span_object_bases(dbg, bank_base + sid, pbytes)
        if len(starts) != 1:
            raise DecodeError(
                f"event root {root.index} header/DBG mapping bank 0x{bank_base:X}, "
                f"storage 0x{sid:X}/{pbytes}B matches {len(starts)} waveform allocations"
            )


def _reconstruct_stream(stream: Stream, wdb: WdbInfo) -> None:
    if stream.payload_bytes <= 0:
        raise DecodeError(f"invalid stream storage size {stream.payload_bytes}")
    kind = stream.storage_kind
    if kind in {"sv_digital", "sv_unpacked_array"} and stream.payload_bytes % 8:
        raise DecodeError(
            f"{kind} allocation 0x{int(stream.absolute_addr or 0):X} has non-aval/bval storage size {stream.payload_bytes}"
        )
    if kind in {"sv_real", "vhdl_real"} and stream.payload_bytes != 8:
        raise DecodeError(f"real allocation 0x{int(stream.absolute_addr or 0):X} is not 8 bytes")

    # Records are normally already chronological because event pages/chunks are
    # read in file order. Avoid an O(N log N) sort for multi-million-event WDBs;
    # sort only if a future file proves otherwise.
    frags = stream.fragments
    prev = None
    ordered = True
    for fr in frags:
        key = (fr.time, fr.chunk_serial, fr.slot)
        if prev is not None and key < prev:
            ordered = False
            break
        prev = key
    if not ordered:
        frags = sorted(frags, key=lambda f: (f.time, f.chunk_serial, f.slot))

    state = bytearray(stream.payload_bytes)
    known = bytearray(stream.payload_bytes)
    complete = False
    events: list[RawEvent] = []
    for fr in frags:
        root_bank = wdb.roots[fr.root_index].bank_base
        if root_bank is None:
            raise DecodeError("internal error: unresolved fragment root bank")
        abs_start = root_bank + fr.storage_id
        off = abs_start - int(stream.absolute_addr or 0)
        if off < 0 or off + len(fr.payload) > stream.payload_bytes:
            raise DecodeError(
                f"fragment root={fr.root_index} sid=0x{fr.storage_id:X}/{len(fr.payload)}B "
                f"falls outside mapped object at 0x{int(stream.absolute_addr or 0):X}"
            )
        if kind in {"sv_digital", "sv_unpacked_array"} and (off % 8 or len(fr.payload) % 8):
            raise DecodeError(
                f"{kind} fragment offset/size {off}/{len(fr.payload)} at runtime address "
                f"0x{int(stream.absolute_addr or 0):X} is not aval/bval-word aligned"
            )
        if kind in {"sv_real", "vhdl_real"} and (off != 0 or len(fr.payload) != 8):
            raise DecodeError(
                f"real fragment at runtime address 0x{int(stream.absolute_addr or 0):X} "
                f"is partial ({off}+{len(fr.payload)} of 8 bytes)"
            )

        state[off:off + len(fr.payload)] = fr.payload
        if not complete:
            known[off:off + len(fr.payload)] = b"\x01" * len(fr.payload)
            complete = 0 not in known
        if complete:
            events.append(RawEvent(fr.time, bytes(state), fr.chunk_serial, fr.slot))

    if not complete:
        try:
            missing = known.index(0)
        except ValueError:
            missing = -1
        raise DecodeError(
            f"runtime address 0x{int(stream.absolute_addr or 0):X} never received a complete "
            f"state snapshot; first missing byte offset={missing}"
        )
    if not events:
        raise DecodeError(
            f"runtime address 0x{int(stream.absolute_addr or 0):X} produced no reconstructable events"
        )
    stream.events = events


def map_roots_to_dbg(wdb: WdbInfo, dbg: DebugDB, wcfg: WcfgInfo | None, log: Logger) -> None:
    if not is_power_of_two(wdb.bank_size):
        raise DecodeError(f"runtime bank size 0x{wdb.bank_size:X} is not a power of two")

    # Build one root index once. This is important for stress WDBs with millions
    # of event fragments.
    fragments_by_root: dict[int, list[RawFragment]] = defaultdict(list)
    for fr in wdb.fragments:
        fragments_by_root[fr.root_index].append(fr)

    log.add("\n=== Runtime bank resolution ===")
    chosen = _header_root_bank_hints(wdb, log)
    if chosen is not None:
        # Header hints are a characterized part of the Event-first WDB 01 layout,
        # but still validate every unique storage shape against DBG/RTTI before
        # trusting them.
        for root, base in zip(wdb.roots, chosen):
            _validate_root_bank(root, base, fragments_by_root[root.index], dbg)
            root.candidate_bases = [base]
            log.add(f"root {root.index:3d}: header-validated candidate=0x{base:X}")
    else:
        candidate_lists: list[list[int]] = []
        bases = sorted({addr & ~(wdb.bank_size - 1) for addr in dbg.objects_by_addr if addr})
        for root in wdb.roots:
            shapes = sorted({(f.storage_id, len(f.payload)) for f in fragments_by_root[root.index]})
            c: list[int] = []
            for bank_base in bases:
                if all(len(_span_object_bases(dbg, bank_base + sid, pbytes)) == 1 for sid, pbytes in shapes):
                    c.append(bank_base)
            root.candidate_bases = c
            candidate_lists.append(c)
            log.add(
                f"root {root.index:3d}: candidates=" + (", ".join(f"0x{x:X}" for x in c) if c else "NONE")
            )
            if not c:
                raise DecodeError(
                    f"event root {root.index} has no DBG runtime-bank candidate; "
                    "record-fragment/type mapping is not supported by this decoder"
                )

        solutions = enumerate_assignments(candidate_lists)
        if not solutions:
            raise DecodeError("no one-to-one event-root/runtime-bank mapping satisfies the DBG database")

        chosen = None
        if len(solutions) == 1:
            chosen = solutions[0]
        elif wcfg:
            wanted_addrs = _selected_dbg_addresses(dbg, wcfg)
            scored: list[tuple[int, tuple[int, ...]]] = []
            for sol in solutions:
                present: set[int] = set()
                for root in wdb.roots:
                    bank = sol[root.index]
                    for fr in fragments_by_root[root.index]:
                        starts = _span_object_bases(dbg, bank + fr.storage_id, len(fr.payload))
                        if len(starts) == 1:
                            present.add(starts[0])
                scored.append((len(present & wanted_addrs), sol))
            best = max(score for score, _ in scored)
            best_solutions = [sol for score, sol in scored if score == best]
            if len(best_solutions) == 1:
                chosen = best_solutions[0]
                log.add(f"WCFG disambiguation    : selected unique mapping score={best}")

        if chosen is None:
            details = "; ".join(
                f"root {i}: " + ",".join(f"0x{x:X}" for x in c)
                for i, c in enumerate(candidate_lists)
            )
            raise DecodeError(
                "DBG mapping is genuinely ambiguous; refusing to guess waveform names. "
                "Supply the matching WCFG or use a WDB within the supported format profile. Candidates: " + details
            )

    assert chosen is not None
    for root, base in zip(wdb.roots, chosen):
        root.bank_base = base
        log.add(f"root {root.index:3d}: resolved bank=0x{base:X}")

    # Map raw WDB fragment records to the unique DBG allocation containing them.
    stream_map: dict[int, Stream] = {}
    for fr in wdb.fragments:
        bank = wdb.roots[fr.root_index].bank_base
        if bank is None:
            raise DecodeError("internal error: unresolved root bank")
        starts = _span_object_bases(dbg, bank + fr.storage_id, len(fr.payload))
        if len(starts) != 1:
            raise DecodeError(
                f"mapped fragment root={fr.root_index} sid=0x{fr.storage_id:X}/{len(fr.payload)}B "
                f"matches {len(starts)} DBG allocations at bank 0x{bank:X}"
            )
        obj_addr = starts[0]
        objs = dbg.objects_by_addr[obj_addr]
        size, _kind = _alias_group_metadata(objs, obj_addr)
        stream = stream_map.get(obj_addr)
        if stream is None:
            stream = Stream(
                root_index=fr.root_index,
                storage_id=obj_addr & (wdb.bank_size - 1),
                payload_bytes=size,
                absolute_addr=obj_addr,
            )
            stream.dbg_objects = sorted(objs, key=lambda o: (o.owner, o.index))
            stream_map[obj_addr] = stream
        stream.fragments.append(fr)

    streams = sorted(stream_map.values(), key=lambda st: (st.root_index, int(st.absolute_addr or 0)))
    for stream in streams:
        _reconstruct_stream(stream, wdb)

    wdb.streams = streams

    log.add("\n=== Exact stream mapping ===")
    for stream in wdb.streams:
        aliases = ", ".join(o.path for o in stream.dbg_objects)
        log.add(
            f"roots={','.join(str(x) for x in sorted({f.root_index for f in stream.fragments})):<7} "
            f"sid=0x{stream.storage_id:04X} addr=0x{int(stream.absolute_addr):05X} "
            f"storage={stream.payload_bytes:4d}B fragments={len(stream.fragments):7d} "
            f"kind={stream.storage_kind:<18} -> {aliases}"
        )

    # Raw fragment objects are no longer needed after complete stream states have
    # been reconstructed. Releasing them here materially reduces peak memory for
    # long simulations with millions of WDB records.
    for stream in wdb.streams:
        stream.fragments.clear()
    wdb.fragments.clear()


# ---------------------------------------------------------------------------
# Snapshot/delta collapse and WCFG selection
# ---------------------------------------------------------------------------

def _canonical_digital_payload(payload: bytes, width: int) -> bytes:
    avals, bvals = digital_payload_words(payload)
    nwords = max(1, (width + 31) // 32)
    if len(avals) != nwords:
        raise DecodeError(
            f"digital payload has {len(avals)} words but declared width {width} needs {nwords}"
        )
    if width % 32:
        mask = (1 << (width % 32)) - 1
        avals[-1] &= mask
        bvals[-1] &= mask
    out = bytearray()
    for a, b in zip(avals, bvals):
        out += struct.pack("<II", a, b)
    return bytes(out)


VHDL_STD_LOGIC_TO_VCD = {
    0: "x",  # U
    1: "x",  # X
    2: "0",
    3: "1",
    4: "z",  # Z
    5: "x",  # W
    6: "0",  # L
    7: "1",  # H
    8: "x",  # - (don't care)
}


def _validate_vhdl_enum_bytes(payload: bytes, allowed_max: int, label: str) -> bytes:
    for i, value in enumerate(payload):
        if value > allowed_max:
            raise DecodeError(
                f"{label} payload contains uncharacterized enumeration ordinal {value} at byte {i}"
            )
    return payload


def canonical_event_payload(stream: Stream, payload: bytes) -> bytes:
    kind = stream.storage_kind
    if kind in {"sv_real", "vhdl_real"}:
        if len(payload) != 8:
            raise DecodeError("real stream payload is not 8 bytes")
        return payload
    if kind == "sv_unpacked_array":
        leaf_real = {bool(o.sv_leaf_type and o.sv_leaf_type.kind == "sv_real") for o in stream.dbg_objects}
        if len(leaf_real) != 1:
            raise DecodeError("unpacked-array aliases disagree on leaf datatype")
        if next(iter(leaf_real)):
            if len(payload) != stream.payload_bytes or len(payload) % 8:
                raise DecodeError("unpacked real-array payload is not a complete sequence of binary64 values")
            return payload
        return _canonical_digital_payload(payload, stream.width)
    if kind == "sv_digital":
        return _canonical_digital_payload(payload, stream.width)
    if kind in {"vhdl_logic", "vhdl_vector"}:
        if len(payload) != stream.payload_bytes:
            raise DecodeError(
                f"{kind} payload has {len(payload)} bytes, expected {stream.payload_bytes}"
            )
        return _validate_vhdl_enum_bytes(payload, 8, kind)
    if kind in {"vhdl_bit", "vhdl_bit_vector", "vhdl_boolean"}:
        if len(payload) != stream.payload_bytes:
            raise DecodeError(
                f"{kind} payload has {len(payload)} bytes, expected {stream.payload_bytes}"
            )
        return _validate_vhdl_enum_bytes(payload, 1, kind)
    if kind == "vhdl_integer":
        if len(payload) != 4:
            raise DecodeError(f"VHDL INTEGER payload is {len(payload)} bytes, expected 4")
        return payload
    if kind == "vhdl_time":
        if len(payload) != 8:
            raise DecodeError(f"VHDL TIME payload is {len(payload)} bytes, expected 8")
        return payload
    if kind == "vhdl_enum":
        if len(payload) != 1:
            raise DecodeError(f"VHDL enumeration payload is {len(payload)} bytes, expected 1")
        counts = {len(o.type_info.enum_literals) for o in stream.dbg_objects if o.type_info}
        if len(counts) != 1 or next(iter(counts)) <= 0:
            raise DecodeError("VHDL enumeration aliases disagree on literal table")
        return _validate_vhdl_enum_bytes(payload, next(iter(counts)) - 1, "vhdl_enum")
    if kind in {"vhdl_record", "vhdl_array"}:
        if len(payload) != stream.payload_bytes:
            raise DecodeError(
                f"{kind} payload has {len(payload)} bytes, expected {stream.payload_bytes}"
            )
        # Composite leaf validation happens while expanding the RTTI-described
        # fields/elements below.  Keeping the raw aggregate here preserves exact
        # snapshot comparison without pretending it is an aval/bval vector.
        return payload
    raise DecodeError(f"unsupported mapped storage kind {kind!r}")


def collapse_stream(stream: Stream) -> None:
    stream.events.sort(key=lambda e: (e.time, e.chunk_serial, e.slot))
    collapsed: list[RawEvent] = []
    canonical_values: list[bytes] = []
    for e in stream.events:
        cp = canonical_event_payload(stream, e.payload)
        if collapsed and collapsed[-1].time == e.time:
            # WDB preserves delta states; XSim's VCD output exposes the settled
            # state at a physical timestamp. Keep the final state at that time.
            collapsed[-1] = e
            canonical_values[-1] = cp
            continue
        if canonical_values and canonical_values[-1] == cp:
            # Periodic chunk-boundary state snapshot, not a value change.
            continue
        collapsed.append(e)
        canonical_values.append(cp)
    stream.events = collapsed


def apply_wcfg_selection(wdb: WdbInfo, dbg: DebugDB, wcfg: WcfgInfo | None, log: Logger) -> list[str]:
    if wcfg is None:
        for s in wdb.streams:
            s.selected_objects = list(s.dbg_objects)
        return []

    wanted = {x.canonical for x in wcfg.signals}
    found: set[str] = set()
    for s in wdb.streams:
        selected = []
        for o in s.dbg_objects:
            cp = canonical_path(o.path)
            if cp in wanted:
                selected.append(o)
                found.add(cp)
        s.selected_objects = selected
    missing = [x.path for x in wcfg.signals if x.canonical not in found]
    log.add("\n=== WCFG selection ===")
    log.add(f"file                  : {wcfg.path}")
    log.add(f"waveform objects      : {len(wcfg.signals)}")
    log.add(f"stored/matched        : {len(found)}")
    log.add(f"not stored in WDB     : {len(missing)}")
    for p in missing[:50]:
        log.add(f"  missing: {p}")
    if len(missing) > 50:
        log.add(f"  ... {len(missing)-50} more")
    return missing


# ---------------------------------------------------------------------------
# WDB payload -> normalized digital text
# ---------------------------------------------------------------------------

def _vhdl_logic_char(value: int) -> str:
    try:
        return VHDL_STD_LOGIC_TO_VCD[value]
    except KeyError as exc:
        raise DecodeError(f"unsupported VHDL std_logic ordinal {value}") from exc


def digital_value_text(stream: Stream, payload: bytes) -> str:
    width = stream.width
    kind = stream.storage_kind

    if kind in {"sv_digital", "sv_unpacked_array"}:
        avals, bvals = digital_payload_words(_canonical_digital_payload(payload, width))
        if width == 1:
            a = avals[0] & 1
            b = bvals[0] & 1
            return "x" if a and b else "z" if b else "1" if a else "0"

        aval = 0
        bval = 0
        for i, (a, b) in enumerate(zip(avals, bvals)):
            aval |= a << (32 * i)
            bval |= b << (32 * i)
        if bval == 0:
            return "b" + format(aval, f"0{width}b")
        chars: list[str] = []
        for bit in range(width - 1, -1, -1):
            a = (aval >> bit) & 1
            b = (bval >> bit) & 1
            chars.append("x" if a and b else "z" if b else "1" if a else "0")
        return "b" + "".join(chars)

    if kind == "vhdl_logic":
        if len(payload) != 1:
            raise DecodeError(f"VHDL std_logic payload is {len(payload)} bytes")
        return _vhdl_logic_char(payload[0])

    if kind == "vhdl_vector":
        if len(payload) != width:
            raise DecodeError(
                f"VHDL vector payload length {len(payload)} does not match declared width {width}"
            )
        # XSim stores one std_logic enumeration ordinal per element, in the
        # declaration's left-to-right range order. That is also the textual VCD
        # vector order, so no byte reversal is required.
        return "b" + "".join(_vhdl_logic_char(v) for v in payload)

    if kind == "vhdl_bit":
        if len(payload) != 1 or payload[0] > 1:
            raise DecodeError("invalid VHDL BIT payload")
        return "1" if payload[0] else "0"

    if kind == "vhdl_bit_vector":
        if len(payload) != width or any(v > 1 for v in payload):
            raise DecodeError("invalid VHDL BIT_VECTOR payload")
        return "b" + "".join("1" if v else "0" for v in payload)

    if kind == "vhdl_boolean":
        if len(payload) != 1 or payload[0] > 1:
            raise DecodeError("invalid VHDL BOOLEAN payload")
        return "1" if payload[0] else "0"

    if kind == "vhdl_integer":
        if len(payload) != 4:
            raise DecodeError("VHDL INTEGER storage is not 4 bytes")
        raw = int.from_bytes(payload, "little", signed=False)
        return "b" + format(raw, "032b")

    if kind == "vhdl_time":
        if len(payload) != 8:
            raise DecodeError("VHDL TIME storage is not 8 bytes")
        raw = int.from_bytes(payload, "little", signed=False)
        return "b" + format(raw, "064b")

    if kind == "vhdl_enum":
        if len(payload) != 1:
            raise DecodeError("VHDL enumeration storage is not 1 byte")
        counts = {len(o.type_info.enum_literals) for o in stream.dbg_objects if o.type_info}
        if len(counts) != 1 or next(iter(counts)) <= 0 or payload[0] >= next(iter(counts)):
            raise DecodeError("invalid VHDL enumeration ordinal")
        return "b" + format(payload[0], f"0{width}b")

    raise DecodeError(f"storage kind {kind!r} is not a digital waveform representation")


def _vhdl_leaf_value(type_info: RttiType, payload: bytes, width: int) -> str | float:
    """Decode one characterized VHDL scalar/vector leaf from native WDB storage."""
    kind = type_info.kind
    if kind == "vhdl_real":
        if len(payload) != 8:
            raise DecodeError("VHDL REAL leaf is not 8 bytes")
        return struct.unpack("<d", payload)[0]
    if kind == "vhdl_logic":
        if len(payload) != 1:
            raise DecodeError("VHDL logic leaf is not 1 byte")
        return _vhdl_logic_char(payload[0])
    if kind == "vhdl_vector":
        if len(payload) != width:
            raise DecodeError(f"VHDL vector leaf has {len(payload)} bytes, expected {width}")
        return "".join(_vhdl_logic_char(v) for v in payload)
    if kind == "vhdl_bit":
        if len(payload) != 1 or payload[0] > 1:
            raise DecodeError("invalid VHDL BIT leaf")
        return "1" if payload[0] else "0"
    if kind == "vhdl_bit_vector":
        if len(payload) != width or any(v > 1 for v in payload):
            raise DecodeError("invalid VHDL BIT_VECTOR leaf")
        return "".join("1" if v else "0" for v in payload)
    if kind == "vhdl_boolean":
        if len(payload) != 1 or payload[0] > 1:
            raise DecodeError("invalid VHDL BOOLEAN leaf")
        return "1" if payload[0] else "0"
    if kind == "vhdl_integer":
        if len(payload) != 4:
            raise DecodeError("VHDL INTEGER leaf is not 4 bytes")
        return format(int.from_bytes(payload, "little", signed=False), "032b")
    if kind == "vhdl_time":
        if len(payload) != 8:
            raise DecodeError("VHDL TIME leaf is not 8 bytes")
        return format(int.from_bytes(payload, "little", signed=False), "064b")
    if kind == "vhdl_enum":
        if len(payload) != 1 or not type_info.enum_literals or payload[0] >= len(type_info.enum_literals):
            raise DecodeError(f"invalid VHDL enumeration leaf for {type_info.name!r}")
        ew = max(1, (len(type_info.enum_literals) - 1).bit_length())
        return format(payload[0], f"0{ew}b")
    raise DecodeError(f"uncharacterized VHDL leaf type {type_info.name!r}/{kind}")


def _vhdl_leaf_var_type(type_info: RttiType) -> str:
    if type_info.kind == "vhdl_real":
        return "real"
    if type_info.kind == "vhdl_integer":
        return "integer"
    if type_info.kind == "vhdl_time":
        return "time"
    return "wire"


def _vhdl_leaf_signed(type_info: RttiType) -> bool:
    uname = type_info.name.upper()
    return type_info.kind == "vhdl_integer" or uname in {"SIGNED", "UNRESOLVED_SIGNED"}


def _triplet_range_count(r: tuple[int, int, int]) -> int:
    left, right, step = r
    if step == 1 and right >= left:
        return right - left + 1
    if step == -1 and left >= right:
        return left - right + 1
    raise DecodeError(f"invalid VHDL RTTI range ({left},{right},{step})")


def _vhdl_record_field_spec(field: RttiField, types: list[RttiType]) -> tuple[RttiType, int, int, bool, int | None, int | None]:
    """Return (type,width,payload_bytes,is_real,left,right) for a record field."""
    if field.type_index >= len(types):
        raise DecodeError(f"VHDL record field {field.name!r} has invalid type index {field.type_index}")
    t = types[field.type_index]
    if t.kind in {"vhdl_vector", "vhdl_bit_vector"}:
        if not field.ranges:
            raise DecodeError(f"VHDL record vector field {field.name!r} has no concrete range")
        width = 1
        for r in field.ranges:
            width *= _triplet_range_count(r)
        left, right, _step = field.ranges[0]
        return t, width, width, False, left, right
    if field.ranges:
        raise DecodeError(
            f"VHDL record scalar field {field.name!r}/{t.kind} unexpectedly has range metadata"
        )
    if t.kind in {"vhdl_logic", "vhdl_bit", "vhdl_boolean"}:
        return t, 1, 1, False, None, None
    if t.kind == "vhdl_enum":
        width = max(1, (len(t.enum_literals)-1).bit_length())
        return t, width, 1, False, None, None
    if t.kind == "vhdl_integer":
        return t, 32, 4, False, None, None
    if t.kind == "vhdl_time":
        return t, 64, 8, False, None, None
    if t.kind == "vhdl_real":
        return t, 64, 8, True, None, None
    raise DecodeError(
        f"VHDL record field {field.name!r} uses uncharacterized nested type {t.name!r}/{t.kind}"
    )


# ===========================================================================
# Unified waveform model / VCD parser / exporters / GUI / CLI
# ===========================================================================

SCRIPT_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = SCRIPT_DIR / "converted_output"

TIME_UNIT_FS = {
    "fs": 1,
    "ps": 1_000,
    "ns": 1_000_000,
    "us": 1_000_000_000,
    "ms": 1_000_000_000_000,
    "s": 1_000_000_000_000_000,
}


class ConversionError(RuntimeError):
    pass


@dataclass(slots=True)
class Change:
    time: int
    value: str | float


@dataclass(slots=True)
class SignalInfo:
    key: str
    name: str                 # user-facing dot hierarchy
    hdl_path: str             # canonical /hierarchy/path
    scopes: tuple[str, ...]
    ref_name: str
    width: int
    var_type: str
    stream_key: str
    is_real: bool = False
    signed: bool = False
    style: str | None = None
    color: str | None = None
    wcfg_radix: str | None = None
    range_left: int | None = None
    range_right: int | None = None


@dataclass(slots=True)
class StreamInfo:
    key: str
    changes: list[Change] = field(default_factory=list)
    signal_keys: list[str] = field(default_factory=list)


@dataclass(slots=True)
class Waveform:
    source: Path
    source_kind: str
    tick_fs: int
    timescale_text: str
    signals: dict[str, SignalInfo]
    signal_order: list[str]
    streams: dict[str, StreamInfo]
    metadata: dict[str, Any] = field(default_factory=dict)
    diagnostic_log: str = ""
    missing_wcfg: list[str] = field(default_factory=list)
    default_selected: set[str] | None = None


# ---------------------------------------------------------------------------
# General helpers
# ---------------------------------------------------------------------------

def ensure_output_dir() -> Path:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    return OUTPUT_DIR


def canonical_hdl_path(path: str) -> str:
    if not path:
        return "/"
    if path.startswith("/"):
        return canonical_path(path)
    return "/" + "/".join(part for part in path.split(".") if part)


def display_name_from_path(path: str) -> str:
    return canonical_hdl_path(path).lstrip("/").replace("/", ".")


def timescale_to_fs(text: str) -> int:
    m = re.fullmatch(r"\s*(1|10|100)\s*(fs|ps|ns|us|ms|s)\s*", text, re.I)
    if not m:
        raise ConversionError(f"unsupported VCD timescale: {text!r}")
    return int(m.group(1)) * TIME_UNIT_FS[m.group(2).lower()]


def parse_vcd_timescale(header: str) -> tuple[str, int]:
    m = re.search(r"\$timescale\s+(.+?)\s+\$end", header, re.I | re.S)
    if not m:
        return "1ps", TIME_UNIT_FS["ps"]
    raw = " ".join(m.group(1).split())
    raw = raw.replace(" ", "")
    return raw, timescale_to_fs(raw)


def default_output_path(source: Path, fmt: str) -> Path:
    ext = {"vcd": ".vcd", "csv": ".csv", "json": ".json", "excel": ".xlsx"}[fmt]
    ensure_output_dir()
    # Avoid overwriting a VCD input when normalizing VCD -> VCD.
    suffix = "_converted" if source.suffix.lower() == ext else ""
    return OUTPUT_DIR / f"{source.stem}{suffix}{ext}"


def normalize_bits(bits: str, width: int) -> str:
    bits = bits.lower()
    if width <= 1:
        return bits[-1:] if bits else "x"
    if len(bits) > width:
        return bits[-width:]
    if len(bits) < width:
        pad = bits[0] if bits and bits[0] in "xz" else "0"
        bits = pad * (width - len(bits)) + bits
    return bits


def collapse_changes(changes: list[Change]) -> list[Change]:
    if not changes:
        return []
    changes.sort(key=lambda c: c.time)
    out: list[Change] = []
    for ch in changes:
        if out and out[-1].time == ch.time:
            out[-1] = ch  # settled value at that timestamp
        elif out and values_equal(out[-1].value, ch.value):
            continue
        else:
            out.append(ch)
    return out


def values_equal(a: str | float, b: str | float) -> bool:
    if isinstance(a, float) and isinstance(b, float):
        if math.isnan(a) and math.isnan(b):
            return True
        if a == 0.0 and b == 0.0:
            return math.copysign(1.0, a) == math.copysign(1.0, b)
        return a == b
    return a == b


def filter_waveform(wf: Waveform, include: list[str] | None = None,
                    exclude: list[str] | None = None,
                    selected: set[str] | None = None,
                    primary_only: bool = False) -> Waveform:
    keep: list[str] = []
    seen_streams: set[str] = set()
    for key in wf.signal_order:
        sig = wf.signals[key]
        names = (sig.name, sig.hdl_path)
        if selected is not None and sig.name not in selected and sig.hdl_path not in selected:
            continue
        if include and not any(fnmatch.fnmatch(n, pat) for pat in include for n in names):
            continue
        if exclude and any(fnmatch.fnmatch(n, pat) for pat in exclude for n in names):
            continue
        if primary_only and sig.stream_key in seen_streams:
            continue
        seen_streams.add(sig.stream_key)
        keep.append(key)

    if not keep:
        raise ConversionError("no signals remain after filtering")

    keep_set = set(keep)
    signals = {k: wf.signals[k] for k in keep}
    stream_keys = {signals[k].stream_key for k in keep}
    streams: dict[str, StreamInfo] = {}
    for sk in stream_keys:
        orig = wf.streams[sk]
        streams[sk] = StreamInfo(sk, orig.changes, [k for k in orig.signal_keys if k in keep_set])
    return Waveform(
        source=wf.source,
        source_kind=wf.source_kind,
        tick_fs=wf.tick_fs,
        timescale_text=wf.timescale_text,
        signals=signals,
        signal_order=keep,
        streams=streams,
        metadata=dict(wf.metadata),
        diagnostic_log=wf.diagnostic_log,
        missing_wcfg=list(wf.missing_wcfg),
        default_selected=(None if wf.default_selected is None else set(wf.default_selected) & {signals[k].name for k in keep}),
    )


# ---------------------------------------------------------------------------
# VCD input
# ---------------------------------------------------------------------------

def _parse_scope_line(line: str) -> str | None:
    m = re.match(r"\$scope\s+\S+\s+(.+?)\s+\$end\s*$", line)
    return m.group(1).strip() if m else None


def _parse_var_line(line: str) -> tuple[str, int, str, str, int | None, int | None] | None:
    m = re.match(r"\$var\s+(\S+)\s+(\d+)\s+(\S+)\s+(.+?)\s+\$end\s*$", line)
    if not m:
        return None
    typ, width_s, vcd_id, ref = m.groups()
    width = int(width_s)
    range_match = re.search(r"\s+\[(-?\d+)\s*:\s*(-?\d+)\]\s*$", ref)
    range_left = range_right = None
    if range_match:
        msb, lsb = map(int, range_match.groups())
        range_left, range_right = msb, lsb
        range_width = abs(msb - lsb) + 1
        if width == 0:
            # XSim can emit parameter declarations as `$var parameter 0 ... [31:0]`.
            # The range is the useful declared width in that case.
            width = range_width
        ref = ref[:range_match.start()].rstrip()
    else:
        ref = re.sub(r"\s+\[[^\]]+\]\s*$", "", ref).rstrip()
    if ref.startswith("\\"):
        # VCD escaped identifier: the leading backslash is syntax, not HDL name data.
        ref = ref[1:].rstrip()
    return typ, width, vcd_id, ref, range_left, range_right


def load_vcd(path: Path) -> Waveform:
    path = path.resolve()
    try:
        f = path.open("r", encoding="utf-8", errors="replace")
    except OSError as exc:
        raise ConversionError(str(exc)) from exc

    with f:
        header_lines: list[str] = []
        scope_stack: list[str] = []
        signals: dict[str, SignalInfo] = {}
        signal_order: list[str] = []
        streams: dict[str, StreamInfo] = {}
        vcd_id_to_stream: dict[str, str] = {}
        vcd_id_width: dict[str, int] = {}
        alias_count: dict[str, int] = {}

        for raw in f:
            line = raw.strip()
            header_lines.append(raw)
            scope = _parse_scope_line(line)
            if scope is not None:
                scope_stack.append(scope)
                continue
            if line == "$upscope $end":
                if scope_stack:
                    scope_stack.pop()
                continue
            var = _parse_var_line(line)
            if var:
                typ, width, vcd_id, ref, range_left, range_right = var
                stream_key = vcd_id_to_stream.setdefault(vcd_id, f"vcd:{vcd_id}")
                streams.setdefault(stream_key, StreamInfo(stream_key))
                vcd_id_width.setdefault(vcd_id, width)
                alias_n = alias_count.get(vcd_id, 0)
                alias_count[vcd_id] = alias_n + 1
                key = f"{stream_key}:{alias_n}"
                hdl_path = "/" + "/".join([*scope_stack, ref])
                sig = SignalInfo(
                    key=key,
                    name=".".join([*scope_stack, ref]),
                    hdl_path=hdl_path,
                    scopes=tuple(scope_stack),
                    ref_name=ref,
                    width=max(1, width),
                    var_type=typ,
                    stream_key=stream_key,
                    is_real=(typ == "real"),
                    signed=(typ == "integer"),
                    range_left=range_left,
                    range_right=range_right,
                )
                signals[key] = sig
                signal_order.append(key)
                streams[stream_key].signal_keys.append(key)
                continue
            if "$enddefinitions" in line:
                break

        header = "".join(header_lines)
        timescale_text, tick_fs = parse_vcd_timescale(header)
        date_match = re.search(r"\$date\s+(.+?)\s+\$end", header, re.S)
        version_match = re.search(r"\$version\s+(.+?)\s+\$end", header, re.S)

        current_time = 0
        for raw in f:
            line = raw.strip()
            if not line or line.startswith("$"):
                continue
            if line.startswith("#"):
                try:
                    current_time = int(line[1:].strip())
                except ValueError:
                    pass
                continue
            if line.startswith("b"):
                m = re.match(r"b([01xXzZ]+)\s+(\S+)", line)
                if not m:
                    continue
                bits, vcd_id = m.groups()
                sk = vcd_id_to_stream.get(vcd_id)
                if sk:
                    width = vcd_id_width.get(vcd_id, len(bits))
                    streams[sk].changes.append(Change(current_time, normalize_bits(bits, width)))
                continue
            if line.startswith("r"):
                m = re.match(r"r([^\s]+)\s+(\S+)", line)
                if not m:
                    continue
                val_s, vcd_id = m.groups()
                sk = vcd_id_to_stream.get(vcd_id)
                if sk:
                    low = val_s.lower()
                    if low in {"nan", "+nan", "-nan"}:
                        val = float("nan")
                    elif low in {"inf", "+inf", "infinity", "+infinity"}:
                        val = float("inf")
                    elif low in {"-inf", "-infinity"}:
                        val = float("-inf")
                    else:
                        try:
                            val = float(val_s)
                        except ValueError:
                            continue
                    streams[sk].changes.append(Change(current_time, val))
                continue
            if line[0] in "01xXzZ" and len(line) >= 2:
                vcd_id = line[1:].strip()
                sk = vcd_id_to_stream.get(vcd_id)
                if sk:
                    streams[sk].changes.append(Change(current_time, line[0].lower()))

    for stream in streams.values():
        stream.changes = collapse_changes(stream.changes)

    return Waveform(
        source=path,
        source_kind="vcd",
        tick_fs=tick_fs,
        timescale_text=timescale_text,
        signals=signals,
        signal_order=signal_order,
        streams=streams,
        metadata={
            "filename": path.name,
            "date": date_match.group(1).strip() if date_match else "",
            "version": version_match.group(1).strip() if version_match else "",
            "input_format": "VCD",
            "timescale": timescale_text,
        },
    )


# ---------------------------------------------------------------------------
# WDB/WCFG input
# ---------------------------------------------------------------------------

def _resolve_inputs(inputs: list[Path], explicit_wcfg: Path | None) -> tuple[Path, WcfgInfo | None]:
    if not inputs:
        raise DecodeError("no input file supplied")
    if len(inputs) > 2:
        raise DecodeError("pass one WDB, one WCFG, or a WDB+WCFG pair")
    wdb: Path | None = None
    wcfg_path: Path | None = explicit_wcfg
    for p0 in inputs:
        p = p0.resolve()
        ext = p.suffix.lower()
        if ext == ".wdb":
            if wdb and wdb != p:
                raise DecodeError("more than one WDB was supplied")
            wdb = p
        elif ext == ".wcfg":
            if wcfg_path and wcfg_path.resolve() != p:
                raise DecodeError("more than one WCFG was supplied")
            wcfg_path = p
        else:
            raise DecodeError(f"unsupported input extension: {p.name}")
    wcfg = parse_wcfg(wcfg_path) if wcfg_path else None
    if wdb is None:
        if not wcfg:
            raise DecodeError("no WDB supplied")
        wdb = resolve_wdb_from_wcfg(wcfg)
    if not wdb.exists():
        raise DecodeError(f"WDB not found: {wdb}")
    return wdb, wcfg


def _resolve_wdb_inputs(inputs: list[Path], explicit_wcfg: Path | None = None) -> tuple[Path, WcfgInfo | None]:
    return _resolve_inputs(inputs, explicit_wcfg)


def load_wdb(inputs: list[Path], explicit_wcfg: Path | None = None) -> Waveform:
    logger = Logger()
    source_hint = (inputs[0] if inputs else explicit_wcfg)
    try:
        wdb_path, wcfg = _resolve_wdb_inputs(inputs, explicit_wcfg)
        info = parse_wdb(wdb_path, logger)
        dbg = parse_debug_db(info.data, logger)
        map_roots_to_dbg(info, dbg, wcfg, logger)

        logger.add("\n=== Stream reconstruction / collapse ===")
        for s in info.streams:
            reconstructed_n = len(s.events)
            collapse_stream(s)
            aliases = ", ".join(o.path for o in s.dbg_objects)
            logger.add(
                f"addr=0x{int(s.absolute_addr):05X} fragments={len(s.fragments):8d} "
                f"reconstructed={reconstructed_n:8d} collapsed={len(s.events):8d} "
                f"width={s.width:<4d} {'real' if s.is_real else 'digital':7s} {aliases}"
            )

        # WCFG is display/selection metadata only. Names, hierarchy, storage,
        # widths and datatypes come from WDB+DBG+RTTI.
        wcfg_selected = {x.canonical for x in wcfg.signals} if wcfg else set()
        style_map: dict[str, WcfgSignal] = {x.canonical: x for x in wcfg.signals} if wcfg else {}

        signals: dict[str, SignalInfo] = {}
        signal_order: list[str] = []
        streams: dict[str, StreamInfo] = {}
        default_selected: set[str] | None = set() if wcfg else None
        generated_paths: set[str] = set()

        active_streams = [s for s in info.streams if s.events and s.dbg_objects]
        for s in active_streams:
            base_sk = f"wdb:{int(s.absolute_addr or 0):X}"

            # VHDL custom arrays are native one-element-per-storage-item
            # aggregates, not Verilog aval/bval vectors.  RTTI gives the outer
            # rank/element type and DBG gives concrete ranges, so expose every
            # aggregate leaf as an ordinary portable waveform signal.
            if s.storage_kind == "vhdl_array":
                arrays = [o for o in s.dbg_objects if o.type_info and o.type_info.kind == "vhdl_array"]
                if len(arrays) != len(s.dbg_objects):
                    raise ConversionError(
                        f"runtime address 0x{int(s.absolute_addr or 0):X} mixes VHDL-array and non-array aliases"
                    )
                shapes = [o.vhdl_array_leaf_shape for o in arrays]
                if any(sh != shapes[0] for sh in shapes[1:]):
                    raise ConversionError(
                        f"runtime address 0x{int(s.absolute_addr or 0):X} has VHDL-array aliases with different shapes"
                    )
                shape = shapes[0]
                if shape is None:
                    raise ConversionError("internal error: VHDL array has no leaf shape")
                index_tuples, leaf_type, leaf_width, leaf_payload, leaf_dims = shape
                if leaf_payload * len(index_tuples) != s.payload_bytes:
                    raise ConversionError(
                        f"VHDL array at 0x{int(s.absolute_addr or 0):X} derives "
                        f"{len(index_tuples)}x{leaf_payload}B leaves but runtime storage is {s.payload_bytes}B"
                    )

                for pos, idx_tuple in enumerate(index_tuples):
                    elem_sk = f"{base_sk}:vhdl_elem:{pos}"
                    off = pos * leaf_payload
                    elem_changes = [
                        Change(e.time, _vhdl_leaf_value(leaf_type, e.payload[off:off + leaf_payload], leaf_width))
                        for e in s.events
                    ]
                    elem_changes = collapse_changes(elem_changes)
                    estream = StreamInfo(elem_sk, elem_changes, [])
                    streams[elem_sk] = estream
                    suffix = "".join(f"[{idx}]" for idx in idx_tuple)
                    for oi, obj in enumerate(s.dbg_objects):
                        cp = canonical_path(obj.path + suffix)
                        parent_cp = canonical_path(obj.path)
                        generated_paths.add(parent_cp)
                        generated_paths.add(cp)
                        wcfg_sig = style_map.get(cp) or style_map.get(parent_cp)
                        key = f"{elem_sk}:{oi}"
                        display_name = display_name_from_path(cp)
                        if leaf_type.kind in {"vhdl_vector", "vhdl_bit_vector"} and len(leaf_dims) == 1:
                            range_left = leaf_dims[0].left
                            range_right = leaf_dims[0].left + leaf_dims[0].step * (leaf_dims[0].count - 1)
                        elif leaf_width > 1 and leaf_type.kind != "vhdl_real":
                            range_left, range_right = leaf_width - 1, 0
                        else:
                            range_left = range_right = None
                        sig = SignalInfo(
                            key=key,
                            name=display_name,
                            hdl_path=cp,
                            scopes=obj.scopes,
                            ref_name=obj.name + suffix,
                            width=leaf_width,
                            var_type=_vhdl_leaf_var_type(leaf_type),
                            stream_key=elem_sk,
                            is_real=(leaf_type.kind == "vhdl_real"),
                            signed=_vhdl_leaf_signed(leaf_type),
                            style=wcfg_sig.style if wcfg_sig else None,
                            color=wcfg_sig.color if wcfg_sig else None,
                            wcfg_radix=wcfg_sig.radix if wcfg_sig else None,
                            range_left=range_left,
                            range_right=range_right,
                        )
                        signals[key] = sig
                        signal_order.append(key)
                        estream.signal_keys.append(key)
                        if default_selected is not None and (cp in wcfg_selected or parent_cp in wcfg_selected):
                            default_selected.add(display_name)
                continue

            # VHDL records are flattened by their RTTI field descriptors.  The
            # characterized logic/vector record layout is sequential field storage
            # followed by zero padding to the DBG allocation size.  Other nested
            # field storage remains fail-closed until characterized.
            if s.storage_kind == "vhdl_record":
                records = [o for o in s.dbg_objects if o.type_info and o.type_info.kind == "vhdl_record"]
                if len(records) != len(s.dbg_objects):
                    raise ConversionError(
                        f"runtime address 0x{int(s.absolute_addr or 0):X} mixes VHDL-record and non-record aliases"
                    )
                schemas = [o.type_info.record_fields for o in records if o.type_info]
                if not schemas or any(sc != schemas[0] for sc in schemas[1:]):
                    raise ConversionError(
                        f"runtime address 0x{int(s.absolute_addr or 0):X} has inconsistent VHDL record field schemas"
                    )
                fields = schemas[0]
                offset = 0
                field_specs: list[tuple[RttiField, RttiType, int, int, bool, int | None, int | None, int]] = []
                for field in fields:
                    ft, fw, fpb, freal, fl, fr = _vhdl_record_field_spec(field, dbg.rtti_types)
                    if ft.kind not in {"vhdl_logic", "vhdl_vector", "vhdl_bit", "vhdl_bit_vector", "vhdl_boolean", "vhdl_enum"}:
                        raise ConversionError(
                            f"VHDL record {records[0].path} field {field.name!r} uses {ft.kind}; "
                            "native alignment for that field class is not yet characterized"
                        )
                    field_specs.append((field, ft, fw, fpb, freal, fl, fr, offset))
                    offset += fpb
                if offset > s.payload_bytes:
                    raise ConversionError(
                        f"VHDL record {records[0].path} fields require {offset}B but storage is {s.payload_bytes}B"
                    )
                if offset < s.payload_bytes:
                    for e in s.events:
                        if any(e.payload[offset:]):
                            raise ConversionError(
                                f"VHDL record {records[0].path} has non-zero bytes in an uncharacterized padding region"
                            )

                for fi, (field, ft, fw, fpb, freal, fl, fr, foff) in enumerate(field_specs):
                    fsk = f"{base_sk}:field:{fi}"
                    fchanges = [
                        Change(e.time, _vhdl_leaf_value(ft, e.payload[foff:foff + fpb], fw))
                        for e in s.events
                    ]
                    fchanges = collapse_changes(fchanges)
                    fstream = StreamInfo(fsk, fchanges, [])
                    streams[fsk] = fstream
                    for oi, obj in enumerate(s.dbg_objects):
                        parent_cp = canonical_path(obj.path)
                        cp = canonical_path(obj.path + "." + field.name)
                        generated_paths.add(parent_cp)
                        generated_paths.add(cp)
                        wcfg_sig = style_map.get(cp) or style_map.get(parent_cp)
                        key = f"{fsk}:{oi}"
                        display_name = display_name_from_path(cp)
                        sig = SignalInfo(
                            key=key,
                            name=display_name,
                            hdl_path=cp,
                            scopes=obj.scopes,
                            ref_name=obj.name + "." + field.name,
                            width=fw,
                            var_type=_vhdl_leaf_var_type(ft),
                            stream_key=fsk,
                            is_real=freal,
                            signed=_vhdl_leaf_signed(ft),
                            style=wcfg_sig.style if wcfg_sig else None,
                            color=wcfg_sig.color if wcfg_sig else None,
                            wcfg_radix=wcfg_sig.radix if wcfg_sig else None,
                            range_left=fl,
                            range_right=fr,
                        )
                        signals[key] = sig
                        signal_order.append(key)
                        fstream.signal_keys.append(key)
                        if default_selected is not None and (cp in wcfg_selected or parent_cp in wcfg_selected):
                            default_selected.add(display_name)
                continue

            # All aliases at one runtime allocation have already been checked
            # for identical storage metadata.  Unpacked arrays are expanded to
            # leaf elements using the recursive RTTI array chain, not by guessing
            # from aggregate width.  This covers multidimensional arrays and
            # arrays of SystemVerilog real values.
            unpacked = [o for o in s.dbg_objects if o.is_unpacked_array]
            if unpacked:
                if len(unpacked) != len(s.dbg_objects):
                    raise ConversionError(
                        f"runtime address 0x{int(s.absolute_addr or 0):X} mixes unpacked-array "
                        "and non-array aliases"
                    )
                shapes = [o.unpacked_leaf_shape for o in unpacked]
                if any(sh != shapes[0] for sh in shapes[1:]):
                    raise ConversionError(
                        f"runtime address 0x{int(s.absolute_addr or 0):X} has aliases with "
                        "different unpacked-array shapes"
                    )
                shape = shapes[0]
                if shape is None:
                    raise ConversionError("internal error: unpacked array has no leaf shape")
                index_tuples, elem_width, leaf_is_real, packed_dims = shape

                aggregate_bits: list[Change] | None = None
                if not leaf_is_real:
                    aggregate_bits = []
                    for e in s.events:
                        txt = digital_value_text(s, e.payload)
                        aggregate_bits.append(Change(e.time, txt[1:] if txt.startswith("b") else txt))

                for pos, idx_tuple in enumerate(index_tuples):
                    elem_sk = f"{base_sk}:elem:{pos}"
                    if leaf_is_real:
                        # Unpacked aggregate storage is low-address/rightmost-declared.
                        # Digital aggregate rendering already reverses that storage into
                        # declaration order; raw binary64 elements need the equivalent
                        # reversal explicitly.
                        storage_pos = len(index_tuples) - 1 - pos
                        elem_changes = [
                            Change(e.time, struct.unpack_from("<d", e.payload, storage_pos * 8)[0])
                            for e in s.events
                        ]
                    else:
                        assert aggregate_bits is not None
                        lo = pos * elem_width
                        hi = lo + elem_width
                        elem_changes = [
                            Change(ch.time, normalize_bits(str(ch.value), s.width)[lo:hi])
                            for ch in aggregate_bits
                        ]
                    elem_changes = collapse_changes(elem_changes)
                    estream = StreamInfo(elem_sk, elem_changes, [])
                    streams[elem_sk] = estream

                    suffix = "".join(f"[{idx}]" for idx in idx_tuple)
                    for oi, obj in enumerate(s.dbg_objects):
                        cp = canonical_path(obj.path + suffix)
                        parent_cp = canonical_path(obj.path)
                        generated_paths.add(parent_cp)
                        generated_paths.add(cp)
                        wcfg_sig = style_map.get(cp) or style_map.get(parent_cp)
                        key = f"{elem_sk}:{oi}"
                        display_name = display_name_from_path(cp)
                        ref_name = obj.name + suffix
                        if leaf_is_real:
                            range_left = range_right = None
                        elif len(packed_dims) == 1 and packed_dims[0].count == elem_width:
                            range_left = packed_dims[0].left
                            range_right = packed_dims[0].left + packed_dims[0].step * (packed_dims[0].count - 1)
                        else:
                            # VCD has only one packed range.  Multiple packed
                            # dimensions are represented as the equivalent flat vector.
                            range_left = elem_width - 1 if elem_width > 1 else None
                            range_right = 0 if elem_width > 1 else None
                        sig = SignalInfo(
                            key=key,
                            name=display_name,
                            hdl_path=cp,
                            scopes=obj.scopes,
                            ref_name=ref_name,
                            width=elem_width,
                            var_type=("real" if leaf_is_real else obj.vcd_var_type),
                            stream_key=elem_sk,
                            is_real=leaf_is_real,
                            signed=(False if leaf_is_real else obj.is_signed),
                            style=wcfg_sig.style if wcfg_sig else None,
                            color=wcfg_sig.color if wcfg_sig else None,
                            wcfg_radix=wcfg_sig.radix if wcfg_sig else None,
                            range_left=range_left,
                            range_right=range_right,
                        )
                        signals[key] = sig
                        signal_order.append(key)
                        estream.signal_keys.append(key)
                        if default_selected is not None and (cp in wcfg_selected or parent_cp in wcfg_selected):
                            default_selected.add(display_name)
                continue

            if s.is_real:
                aggregate_changes = [
                    Change(e.time, struct.unpack("<d", e.payload)[0])
                    for e in s.events
                ]
            else:
                aggregate_changes = []
                for e in s.events:
                    txt = digital_value_text(s, e.payload)
                    aggregate_changes.append(Change(e.time, txt[1:] if txt.startswith("b") else txt))

            # Ordinary scalar/vector/real allocation. Aliases share one event stream.
            sk = base_sk
            stream = StreamInfo(sk, collapse_changes(aggregate_changes), [])
            streams[sk] = stream
            for oi, obj in enumerate(s.dbg_objects):
                cp = canonical_path(obj.path)
                generated_paths.add(cp)
                wcfg_sig = style_map.get(cp)
                key = f"{sk}:{oi}"
                display_name = display_name_from_path(obj.path)
                sig = SignalInfo(
                    key=key,
                    name=display_name,
                    hdl_path=cp,
                    scopes=obj.scopes,
                    ref_name=obj.name,
                    width=obj.bit_width,
                    var_type=obj.vcd_var_type,
                    stream_key=sk,
                    is_real=obj.is_real,
                    signed=obj.is_signed,
                    style=wcfg_sig.style if wcfg_sig else None,
                    color=wcfg_sig.color if wcfg_sig else None,
                    wcfg_radix=wcfg_sig.radix if wcfg_sig else None,
                    range_left=(obj.dimensions[0].left if len(obj.dimensions) == 1 and obj.dimensions[0].count == obj.bit_width else None),
                    range_right=(obj.dimensions[0].left + obj.dimensions[0].step * (obj.dimensions[0].count - 1) if len(obj.dimensions) == 1 and obj.dimensions[0].count == obj.bit_width else None),
                )
                signals[key] = sig
                signal_order.append(key)
                stream.signal_keys.append(key)
                if default_selected is not None and cp in wcfg_selected:
                    default_selected.add(display_name)

        if not signals:
            raise ConversionError("WDB contains no mapped waveform signals")

        missing: list[str] = []
        if wcfg:
            missing = [x.path for x in wcfg.signals if x.canonical not in generated_paths]
            logger.add("\n=== WCFG selection ===")
            logger.add(f"file                  : {wcfg.path}")
            logger.add(f"waveform objects      : {len(wcfg.signals)}")
            logger.add(f"stored/matched        : {len(wcfg.signals)-len(missing)}")
            logger.add(f"not stored in WDB     : {len(missing)}")
            for mp in missing[:50]:
                logger.add(f"  missing: {mp}")
            if len(missing) > 50:
                logger.add(f"  ... {len(missing)-50} more")
            if not default_selected:
                raise ConversionError("none of the WCFG waveform objects are stored in this WDB")

        metadata = {
            "filename": wdb_path.name,
            "input_format": "WDB",
            "wdb_magic": "Xilinx WAVE DATABASE 01",
            "dbg_magic": "Xilinx ISim DBG 006",
            "rtti_magic": "Xilinx ISim TYPE FILE 001",
            "wcfg": str(wcfg.path) if wcfg else None,
            "timescale": "1ps",
            "wdb_roots": len(info.roots),
            "runtime_streams": len(active_streams),
        }
        return Waveform(
            source=wdb_path,
            source_kind="wdb",
            tick_fs=TIME_UNIT_FS["ps"],
            timescale_text="1ps",
            signals=signals,
            signal_order=signal_order,
            streams=streams,
            metadata=metadata,
            diagnostic_log=logger.text(),
            missing_wcfg=missing,
            default_selected=default_selected,
        )
    except Exception as exc:
        if source_hint is not None:
            try:
                hint = Path(source_hint).resolve()
                failure_log = hint.with_suffix(hint.suffix + ".wdbdecode_failure.log")
                textlog = logger.text()
                textlog += "\n=== Decode failure ===\n"
                textlog += f"source                : {hint}\n"
                textlog += f"error                 : {type(exc).__name__}: {exc}\n"
                failure_log.write_text(textlog, encoding="utf-8")
                raise DecodeError(f"{exc}\nDiagnostic log: {failure_log}") from exc
            except DecodeError:
                raise
            except OSError:
                pass
        raise


def load_waveform(inputs: list[Path], explicit_wcfg: Path | None = None) -> Waveform:
    if not inputs:
        raise ConversionError("no input file supplied")
    suffixes = {p.suffix.lower() for p in inputs}
    if ".vcd" in suffixes:
        if len(inputs) != 1 or suffixes != {".vcd"}:
            raise ConversionError("VCD input cannot be mixed with WDB/WCFG inputs")
        return load_vcd(inputs[0])
    if suffixes <= {".wdb", ".wcfg"}:
        return load_wdb(inputs, explicit_wcfg)
    raise ConversionError("input must be .vcd, .wdb, .wcfg, or .wdb + .wcfg")


# ---------------------------------------------------------------------------
# Timeline and value formatting
# ---------------------------------------------------------------------------

def selected_streams(wf: Waveform) -> list[StreamInfo]:
    wanted = {wf.signals[k].stream_key for k in wf.signal_order}
    return [wf.streams[k] for k in wf.streams if k in wanted]


def initial_value_for_stream(wf: Waveform, stream: StreamInfo) -> str | float:
    sig = wf.signals[stream.signal_keys[0]]
    return float("nan") if sig.is_real else "x" * max(1, sig.width)


def iter_timeline(wf: Waveform) -> Iterator[tuple[int, dict[str, str | float]]]:
    streams = selected_streams(wf)
    current: dict[str, str | float] = {s.key: initial_value_for_stream(wf, s) for s in streams}
    heap: list[tuple[int, int, int]] = []
    for si, s in enumerate(streams):
        if s.changes:
            heapq.heappush(heap, (s.changes[0].time, si, 0))

    while heap:
        t = heap[0][0]
        while heap and heap[0][0] == t:
            _, si, ei = heapq.heappop(heap)
            s = streams[si]
            # There can be multiple same-time records in imported VCD; use final.
            current[s.key] = s.changes[ei].value
            ni = ei + 1
            if ni < len(s.changes):
                heapq.heappush(heap, (s.changes[ni].time, si, ni))
        yield t, current.copy()


RADIX_LABELS = {
    "auto": "Auto / WCFG",
    "hex": "Hex",
    "int": "Unsigned",
    "signed": "Signed",
    "bin": "Binary",
    "oct": "Octal",
    "smag": "Sign-Magnitude",
    "real": "Real",
}
RADIX_CHOICES = ("auto", "hex", "int", "signed", "bin", "oct", "smag")


def effective_radix(sig: SignalInfo, default_fmt: str = "auto",
                    radix_rules: list[tuple[str, str]] | None = None) -> str:
    if sig.is_real:
        return "real"
    # Last matching rule wins, making repeated CLI/GUI overrides predictable.
    if radix_rules:
        for pattern, fmt in reversed(radix_rules):
            if fnmatch.fnmatch(sig.name, pattern) or fnmatch.fnmatch(sig.hdl_path, pattern):
                return fmt
    if default_fmt == "auto":
        if sig.wcfg_radix:
            return sig.wcfg_radix
        # A useful fallback when no WCFG display radix exists.
        if sig.var_type == "integer" and sig.signed:
            return "signed"
        return "hex"
    return default_fmt


def format_digital(bits: str, fmt: str, width: int, signed_hint: bool = False) -> str | int:
    bits = normalize_bits(bits, width)
    if any(c in bits for c in "xz"):
        return bits
    if fmt == "bin":
        return bits
    val = int(bits, 2)
    if fmt == "hex":
        digits = max(1, (width + 3) // 4)
        return format(val, f"0{digits}X")
    if fmt == "oct":
        digits = max(1, (width + 2) // 3)
        return format(val, f"0{digits}o")
    if fmt in {"int", "unsigned"}:
        return val
    if fmt == "signed":
        return val - (1 << width) if width and bits[0] == "1" else val
    if fmt == "smag":
        if width <= 1:
            return val
        mag = int(bits[1:], 2)
        return -mag if bits[0] == "1" else mag
    return val


def export_value(sig: SignalInfo, value: str | float, fmt: str) -> str | int | float:
    if sig.is_real:
        if not isinstance(value, float):
            try:
                value = float(value)
            except Exception:
                return str(value)
        if math.isnan(value):
            return "nan"
        if math.isinf(value):
            return "-inf" if value < 0 else "inf"
        return value
    return format_digital(str(value), fmt, sig.width, sig.signed)


def time_as_number(ticks: int, tick_fs: int, unit: str) -> float | int:
    denom = TIME_UNIT_FS[unit]
    num = ticks * tick_fs
    if num % denom == 0:
        return num // denom
    return num / denom


# ---------------------------------------------------------------------------
# Output writers
# ---------------------------------------------------------------------------

def _vcd_id(n: int) -> str:
    chars: list[str] = []
    while True:
        chars.append(chr(33 + n % 94))
        n = n // 94 - 1
        if n < 0:
            break
    return "".join(reversed(chars))


def _vcd_ref_name(name: str) -> str:
    if name.startswith("\\"):
        return name if name.endswith(" ") else name + " "
    if name and (name[0].isalpha() or name[0] in "_$") and all(c.isalnum() or c in "_$" for c in name):
        return name
    return "\\" + name + " "


def _vcd_scope_name(name: str) -> str:
    return name if name and not any(c.isspace() for c in name) else _vcd_ref_name(name)


def _real_vcd_text(value: float) -> str:
    if math.isnan(value):
        return "rnan"
    if math.isinf(value):
        return "r-inf" if value < 0 else "rinf"
    if value == 0.0:
        return "r-0" if math.copysign(1.0, value) < 0 else "r0"
    return "r" + format(value, ".17g")


def _digital_vcd_text(bits: str, width: int) -> str:
    bits = normalize_bits(bits, width)
    if width == 1:
        return bits
    return "b" + bits


def write_vcd(wf: Waveform, output: Path, timescale: str | None = None) -> int:
    timescale = timescale or wf.timescale_text or "1ps"
    out_tick_fs = timescale_to_fs(timescale)
    # All timestamps must be exactly representable in the requested timescale.
    for s in selected_streams(wf):
        for ch in s.changes:
            if (ch.time * wf.tick_fs) % out_tick_fs:
                raise ConversionError(
                    f"timestamp {ch.time} cannot be represented exactly at output timescale {timescale}; "
                    "choose a finer --timescale"
                )

    stream_list = selected_streams(wf)
    stream_ids = {s.key: _vcd_id(i) for i, s in enumerate(stream_list)}
    tree: dict[str, Any] = {}
    for key in wf.signal_order:
        sig = wf.signals[key]
        node = tree
        for scope in sig.scopes:
            node = node.setdefault(scope, {})
        node.setdefault("__signals__", []).append(sig)

    output.parent.mkdir(parents=True, exist_ok=True)
    emitted = 0
    with output.open("w", encoding="utf-8", newline="\n") as f:
        f.write("$date\n    generated by Vivado Waveform Converter\n$end\n")
        f.write("$version\n    waveform_converter.py\n$end\n")
        f.write(f"$timescale {timescale} $end\n")
        if wf.source_kind == "wdb":
            f.write("$comment WDB decoded directly in Python; WCFG display colors/styles are not standard VCD metadata. $end\n")

        def emit_tree(node: dict[str, Any], depth: int) -> None:
            for scope, child in node.items():
                if scope == "__signals__":
                    continue
                f.write("  " * depth + f"$scope module {_vcd_scope_name(scope)} $end\n")
                emit_tree(child, depth + 1)
                f.write("  " * depth + "$upscope $end\n")
            for sig in node.get("__signals__", []):
                vid = stream_ids[sig.stream_key]
                if sig.is_real:
                    # XSim-compatible real declaration.
                    f.write("  " * depth + f"$var real 32 {vid} {_vcd_ref_name(sig.ref_name)} $end\n")
                else:
                    if sig.width > 1:
                        rl = sig.range_left if sig.range_left is not None else sig.width - 1
                        rr = sig.range_right if sig.range_right is not None else 0
                        rng = f" [{rl}:{rr}]"
                    else:
                        rng = ""
                    f.write("  " * depth + f"$var {sig.var_type} {sig.width} {vid} {_vcd_ref_name(sig.ref_name)}{rng} $end\n")

        emit_tree(tree, 0)
        f.write("$enddefinitions $end\n")
        f.write("$dumpvars\n")

        start_idx: dict[str, int] = {}
        for s in stream_list:
            idx = 0
            if s.changes and s.changes[0].time == 0:
                sig = wf.signals[s.signal_keys[0]]
                val = s.changes[0].value
                vid = stream_ids[s.key]
                txt = _real_vcd_text(float(val)) if sig.is_real else _digital_vcd_text(str(val), sig.width)
                f.write(f"{txt} {vid}\n" if sig.is_real or sig.width > 1 else f"{txt}{vid}\n")
                emitted += 1
                idx = 1
            start_idx[s.key] = idx
        f.write("$end\n")

        heap: list[tuple[int, int, int]] = []
        for si, s in enumerate(stream_list):
            ei = start_idx[s.key]
            if ei < len(s.changes):
                heapq.heappush(heap, (s.changes[ei].time, si, ei))
        current_time: int | None = None
        while heap:
            t, si, ei = heapq.heappop(heap)
            s = stream_list[si]
            ch = s.changes[ei]
            out_t = (t * wf.tick_fs) // out_tick_fs
            if current_time != out_t:
                f.write(f"#{out_t}\n")
                current_time = out_t
            sig = wf.signals[s.signal_keys[0]]
            vid = stream_ids[s.key]
            txt = _real_vcd_text(float(ch.value)) if sig.is_real else _digital_vcd_text(str(ch.value), sig.width)
            f.write(f"{txt} {vid}\n" if sig.is_real or sig.width > 1 else f"{txt}{vid}\n")
            emitted += 1
            ni = ei + 1
            if ni < len(s.changes):
                heapq.heappush(heap, (s.changes[ni].time, si, ni))
    return emitted


def write_csv(wf: Waveform, output: Path, time_unit: str, value_fmt: str,
              radix_rules: list[tuple[str, str]] | None = None) -> int:
    output.parent.mkdir(parents=True, exist_ok=True)
    with output.open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow([f"Time({time_unit})"] + [wf.signals[k].name for k in wf.signal_order])
        rows = 0
        for t, current in iter_timeline(wf):
            vals = []
            for key in wf.signal_order:
                sig = wf.signals[key]
                fmt = effective_radix(sig, value_fmt, radix_rules)
                vals.append(export_value(sig, current[sig.stream_key], fmt))
            writer.writerow([time_as_number(t, wf.tick_fs, time_unit)] + vals)
            rows += 1
    return rows


def _json_safe_value(value: str | int | float) -> str | int | float:
    if isinstance(value, float):
        if math.isnan(value):
            return "nan"
        if math.isinf(value):
            return "-inf" if value < 0 else "inf"
    return value


def write_json(wf: Waveform, output: Path, time_unit: str, value_fmt: str, event_mode: bool = False,
               radix_rules: list[tuple[str, str]] | None = None) -> int:
    output.parent.mkdir(parents=True, exist_ok=True)
    sig_desc = []
    for key in wf.signal_order:
        s = wf.signals[key]
        sig_desc.append({
            "id": key,
            "name": s.name,
            "hdl_path": s.hdl_path,
            "width": s.width,
            "type": s.var_type,
            "real": s.is_real,
            "signed": s.signed,
            "alias_group": s.stream_key,
            "wcfg_style": s.style,
            "wcfg_color": s.color,
            "radix": effective_radix(s, value_fmt, radix_rules),
            "wcfg_radix": s.wcfg_radix,
            "range_left": s.range_left,
            "range_right": s.range_right,
        })

    with output.open("w", encoding="utf-8") as f:
        f.write("{\n")
        f.write('  "metadata": ' + json.dumps(wf.metadata, indent=2).replace("\n", "\n  ") + ",\n")
        f.write('  "time_unit": ' + json.dumps(time_unit) + ",\n")
        f.write('  "signals": ' + json.dumps(sig_desc, indent=2).replace("\n", "\n  ") + ",\n")
        f.write('  "mode": ' + json.dumps("events" if event_mode else "timeline") + ",\n")
        f.write('  "data": [\n')
        count = 0
        first = True
        if event_mode:
            # Compact event-oriented JSON. Merge streams with a heap so a long
            # capture does not require a second in-memory copy of every event.
            streams0 = selected_streams(wf)
            heap0: list[tuple[int, int, int]] = []
            for si, stream in enumerate(streams0):
                if stream.changes:
                    heapq.heappush(heap0, (stream.changes[0].time, si, 0))
            while heap0:
                t, si, ei = heapq.heappop(heap0)
                stream = streams0[si]
                ch = stream.changes[ei]
                sig = wf.signals[stream.signal_keys[0]]
                row = {
                    "time": time_as_number(t, wf.tick_fs, time_unit),
                    "signal": sig.name,
                    "value": _json_safe_value(export_value(sig, ch.value, effective_radix(sig, value_fmt, radix_rules))),
                }
                if not first:
                    f.write(",\n")
                f.write("    " + json.dumps(row, ensure_ascii=False))
                first = False
                count += 1
                ni = ei + 1
                if ni < len(stream.changes):
                    heapq.heappush(heap0, (stream.changes[ni].time, si, ni))
        else:
            for t, current in iter_timeline(wf):
                row: dict[str, Any] = {"time": time_as_number(t, wf.tick_fs, time_unit)}
                for key in wf.signal_order:
                    sig = wf.signals[key]
                    row[sig.name] = _json_safe_value(export_value(sig, current[sig.stream_key], effective_radix(sig, value_fmt, radix_rules)))
                if not first:
                    f.write(",\n")
                f.write("    " + json.dumps(row, ensure_ascii=False))
                first = False
                count += 1
        f.write("\n  ]\n}\n")
    return count


def _ensure_openpyxl():
    try:
        import openpyxl  # type: ignore
        return openpyxl
    except ImportError as exc:
        raise ConversionError(
            "Excel export requires openpyxl. Install it with: "
            "python -m pip install openpyxl"
        ) from exc


def write_excel(wf: Waveform, output: Path, time_unit: str, value_fmt: str,
                radix_rules: list[tuple[str, str]] | None = None) -> int:
    openpyxl = _ensure_openpyxl()
    output.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet("Waveform")
    # Remove default sheet if present in write-only implementation is not needed.
    headers = [f"Time({time_unit})"] + [wf.signals[k].name for k in wf.signal_order]
    ws.append(headers)
    MAX_DATA_ROWS = 1_048_575
    sheet_index = 1
    sheet_rows = 0
    total = 0
    for t, current in iter_timeline(wf):
        if sheet_rows >= MAX_DATA_ROWS:
            sheet_index += 1
            ws = wb.create_sheet(f"Waveform_{sheet_index}")
            ws.append(headers)
            sheet_rows = 0
        vals: list[Any] = [time_as_number(t, wf.tick_fs, time_unit)]
        for key in wf.signal_order:
            sig = wf.signals[key]
            fmt = effective_radix(sig, value_fmt, radix_rules)
            vals.append(export_value(sig, current[sig.stream_key], fmt))
        ws.append(vals)
        sheet_rows += 1
        total += 1
    wb.save(output)
    return total


def export_waveform(wf: Waveform, output: Path, fmt: str, time_unit: str = "us",
                    value_fmt: str = "auto", timescale: str | None = None,
                    json_events: bool = False,
                    radix_rules: list[tuple[str, str]] | None = None) -> int:
    if fmt == "vcd":
        if wf.source_kind == "vcd":
            raise ConversionError("VCD to VCD conversion is intentionally not supported")
        return write_vcd(wf, output, timescale)
    if fmt == "csv":
        return write_csv(wf, output, time_unit, value_fmt, radix_rules)
    if fmt == "json":
        return write_json(wf, output, time_unit, value_fmt, json_events, radix_rules)
    if fmt == "excel":
        return write_excel(wf, output, time_unit, value_fmt, radix_rules)
    raise ConversionError(f"unsupported output format: {fmt}")


# ---------------------------------------------------------------------------
# Diagnostics / listing
# ---------------------------------------------------------------------------

def print_signals(wf: Waveform) -> None:
    print(f"\nAvailable signals ({len(wf.signal_order)} declarations):")
    print("-" * 126)
    selected = wf.default_selected
    for i, key in enumerate(wf.signal_order, 1):
        s = wf.signals[key]
        kind = "real" if s.is_real else f"{s.width}-bit"
        alias = ""
        group = wf.streams[s.stream_key].signal_keys
        if len(group) > 1:
            alias = f" aliases={len(group)}"
        mark = "*" if selected is not None and s.name in selected else " "
        radix = RADIX_LABELS.get(effective_radix(s, "auto"), "Real" if s.is_real else "")
        print(f"{mark}{i:4d}. {s.name:<72} {s.var_type:<10} {kind:<10} {radix:<12}{alias}")
    if selected is not None:
        print("  * = selected by WCFG")
    print("-" * 126)


def write_wdb_log_if_needed(wf: Waveform, output: Path) -> Path | None:
    if wf.source_kind != "wdb" or not wf.diagnostic_log:
        return None
    log_path = output.with_suffix(output.suffix + ".wdbdecode.log")
    text = wf.diagnostic_log
    text += "\n=== All-in-one converter ===\n"
    text += f"output                : {output.resolve()}\n"
    if wf.missing_wcfg:
        text += f"WCFG missing objects  : {len(wf.missing_wcfg)}\n"
    log_path.write_text(text, encoding="utf-8")
    return log_path


# ---------------------------------------------------------------------------
# GUI
# ---------------------------------------------------------------------------

def primary_signal_names(wf: Waveform) -> set[str]:
    """Choose one shortest-hierarchy declaration for each runtime stream."""
    chosen: set[str] = set()
    order_index = {key: i for i, key in enumerate(wf.signal_order)}
    for stream in selected_streams(wf):
        keys = [k for k in stream.signal_keys if k in wf.signals]
        if not keys:
            continue
        key = min(keys, key=lambda k: (len(wf.signals[k].scopes), order_index.get(k, 1 << 30)))
        chosen.add(wf.signals[key].name)
    return chosen


def default_gui_selection(wf: Waveform) -> set[str]:
    if wf.default_selected is not None:
        return set(wf.default_selected)
    if wf.source_kind == "wdb":
        return primary_signal_names(wf)
    return {wf.signals[k].name for k in wf.signal_order}


class SignalSelectorDialog:
    """Searchable selector. The main window only displays the resulting selection."""
    def __init__(self, parent: tk.Misc, wf: Waveform, selected: set[str]):
        self.wf = wf
        self.result: set[str] | None = None
        self.working = set(selected)
        self.dialog = tk.Toplevel(parent)
        self.dialog.title("Select signals")
        self.dialog.transient(parent)
        self.dialog.grab_set()
        self.dialog.geometry("880x620")
        self.dialog.minsize(700, 450)

        main = ttk.Frame(self.dialog, padding=12)
        main.pack(fill="both", expand=True)
        top = ttk.Frame(main)
        top.pack(fill="x", pady=(0, 8))
        ttk.Label(top, text="Filter:").pack(side="left")
        self.filter_var = tk.StringVar()
        ent = ttk.Entry(top, textvariable=self.filter_var)
        ent.pack(side="left", fill="x", expand=True, padx=(6, 8))
        ent.focus_set()
        self.count_label = ttk.Label(top)
        self.count_label.pack(side="right")

        buttons = ttk.Frame(main)
        buttons.pack(fill="x", pady=(0, 8))
        ttk.Button(buttons, text="Select all", command=self._select_all).pack(side="left")
        ttk.Button(buttons, text="Clear", command=self._clear).pack(side="left", padx=5)
        ttk.Button(buttons, text="Primary signals", command=self._primary).pack(side="left", padx=5)
        if wf.default_selected is not None:
            ttk.Button(buttons, text="WCFG selection", command=self._wcfg).pack(side="left", padx=5)

        cols = ("use", "signal", "type", "width", "radix")
        self.tree = ttk.Treeview(main, columns=cols, show="headings", selectmode="extended")
        self.tree.heading("use", text="Use")
        self.tree.heading("signal", text="Signal")
        self.tree.heading("type", text="Type")
        self.tree.heading("width", text="Width")
        self.tree.heading("radix", text="WCFG radix")
        self.tree.column("use", width=45, anchor="center", stretch=False)
        self.tree.column("signal", width=510, anchor="w")
        self.tree.column("type", width=90, anchor="w", stretch=False)
        self.tree.column("width", width=65, anchor="center", stretch=False)
        self.tree.column("radix", width=110, anchor="w", stretch=False)
        scroll = ttk.Scrollbar(main, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scroll.set)
        self.tree.pack(side="left", fill="both", expand=True)
        scroll.pack(side="right", fill="y")
        self.tree.bind("<Double-1>", self._toggle_event)
        self.tree.bind("<space>", self._toggle_event)
        self.filter_var.trace_add("write", lambda *_: self._refresh())

        bottom = ttk.Frame(self.dialog, padding=(12, 0, 12, 12))
        bottom.pack(fill="x")
        ttk.Label(bottom, text="Double-click or press Space to include/exclude highlighted signals.", foreground="gray").pack(side="left")
        ttk.Button(bottom, text="Cancel", command=self.dialog.destroy).pack(side="right", padx=(5, 0))
        ttk.Button(bottom, text="Apply", command=self._ok).pack(side="right")
        self._refresh()
        parent.wait_window(self.dialog)

    def _visible_keys(self) -> list[str]:
        q = self.filter_var.get().strip().lower()
        keys = []
        for key in self.wf.signal_order:
            sig = self.wf.signals[key]
            if q and q not in sig.name.lower() and q not in sig.hdl_path.lower():
                continue
            keys.append(key)
        return keys

    def _refresh(self) -> None:
        self.tree.delete(*self.tree.get_children())
        for key in self._visible_keys():
            sig = self.wf.signals[key]
            use = "✓" if sig.name in self.working else ""
            typ = "real" if sig.is_real else sig.var_type
            radix = RADIX_LABELS.get(sig.wcfg_radix or "", "")
            self.tree.insert("", "end", iid=key, values=(use, sig.name, typ, sig.width, radix))
        self.count_label.config(text=f"{len(self.working)} selected / {len(self.wf.signal_order)} available")

    def _toggle_event(self, _event=None) -> str:
        items = self.tree.selection()
        if not items:
            item = self.tree.focus()
            items = (item,) if item else ()
        for key in items:
            if key not in self.wf.signals:
                continue
            name = self.wf.signals[key].name
            if name in self.working:
                self.working.remove(name)
            else:
                self.working.add(name)
        self._refresh()
        return "break"

    def _select_all(self) -> None:
        self.working = {self.wf.signals[k].name for k in self.wf.signal_order}
        self._refresh()

    def _clear(self) -> None:
        self.working.clear()
        self._refresh()

    def _primary(self) -> None:
        self.working = primary_signal_names(self.wf)
        self._refresh()

    def _wcfg(self) -> None:
        self.working = set(self.wf.default_selected or ())
        self._refresh()

    def _ok(self) -> None:
        if not self.working:
            messagebox.showwarning("No signals", "Select at least one signal.", parent=self.dialog)
            return
        self.result = set(self.working)
        self.dialog.destroy()


class ConverterGUI:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title(f"Vivado Waveform Converter {APP_VERSION}")
        self.root.geometry("1040x760")
        self.root.minsize(850, 620)

        self.input_file = tk.StringVar()
        self.wcfg_file = tk.StringVar()
        self.output_file = tk.StringVar()
        self.format_var = tk.StringVar(value="VCD")
        self.value_fmt = tk.StringVar(value="auto")
        self.time_unit = tk.StringVar(value="us")
        self.json_events = tk.BooleanVar(value=False)

        self.wf: Waveform | None = None
        self.selected_names: set[str] = set()
        self.radix_overrides: dict[str, str] = {}
        self.input_paths: list[Path] = []
        self.wcfg_paths: list[Path] = []
        self.batch_jobs: list[BatchJob] = []
        self.busy = False
        self._task_queue: queue.Queue[tuple[str, bool, Any]] = queue.Queue()
        self._pending_after_load: str | None = None
        self._controls: list[Any] = []
        self._readonly_controls: list[Any] = []
        self._normal_state_controls: list[Any] = []
        self._auto_load_after_id: str | None = None
        self._suppress_auto_load = False
        self._loading_signature: tuple[str, str] | None = None
        self._loaded_signature: tuple[str, str] | None = None
        self._reload_after_busy = False
        self._current_load_quiet = False
        self._build()
        # Manual path entry also auto-loads once it names an existing supported
        # source. Browse buttons trigger an immediate load themselves.
        self.input_file.trace_add("write", self._input_text_changed)
        self.wcfg_file.trace_add("write", self._wcfg_text_changed)

    def _build(self) -> None:
        style = ttk.Style()
        style.configure("Action.TButton", padding=(16, 7))

        outer = ttk.Frame(self.root, padding=14)
        outer.pack(fill="both", expand=True)
        outer.columnconfigure(0, weight=1)
        outer.rowconfigure(1, weight=1)

        source = ttk.LabelFrame(outer, text="Source", padding=10)
        source.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        source.columnconfigure(1, weight=1)
        ttk.Label(source, text="Waveform:").grid(row=0, column=0, sticky="w", padx=(0, 8), pady=4)
        self.input_entry = ttk.Entry(source, textvariable=self.input_file)
        self.input_entry.grid(row=0, column=1, sticky="ew", pady=4)
        self._normal_state_controls.append(self.input_entry)
        b = ttk.Button(source, text="Browse…", command=self._browse_input); b.grid(row=0, column=2, padx=(8, 0)); self._controls.append(b)
        ttk.Label(source, text="WCFG:").grid(row=1, column=0, sticky="w", padx=(0, 8), pady=4)
        self.wcfg_entry = ttk.Entry(source, textvariable=self.wcfg_file)
        self.wcfg_entry.grid(row=1, column=1, sticky="ew", pady=4)
        self._normal_state_controls.append(self.wcfg_entry)
        b = ttk.Button(source, text="Browse…", command=self._browse_wcfg); b.grid(row=1, column=2, padx=(8, 0)); self._controls.append(b)
        self.source_hint = ttk.Label(
            source,
            text="Sources load automatically. Browse supports multi-select; matching WCFG files are paired automatically.",
            foreground="gray",
        )
        self.source_hint.grid(row=2, column=1, columnspan=2, sticky="w", pady=(3, 0))

        body = ttk.Panedwindow(outer, orient="vertical")
        body.grid(row=1, column=0, sticky="nsew")

        sig_frame = ttk.LabelFrame(body, text="Selected signals", padding=8)
        body.add(sig_frame, weight=3)
        sig_frame.columnconfigure(0, weight=1)
        sig_frame.rowconfigure(1, weight=1)
        toolbar = ttk.Frame(sig_frame)
        toolbar.grid(row=0, column=0, sticky="ew", pady=(0, 6))
        b = ttk.Button(toolbar, text="Select…", command=self._select_signals); b.pack(side="left"); self._controls.append(b)
        b = ttk.Button(toolbar, text="Remove highlighted", command=self._remove_highlighted); b.pack(side="left", padx=5); self._controls.append(b)
        self.signal_summary = ttk.Label(toolbar, text="No waveform loaded", foreground="gray")
        self.signal_summary.pack(side="left", padx=10)

        cols = ("signal", "type", "width", "radix")
        self.signal_tree = ttk.Treeview(sig_frame, columns=cols, show="headings", selectmode="extended", height=12)
        self.signal_tree.heading("signal", text="Signal")
        self.signal_tree.heading("type", text="Type")
        self.signal_tree.heading("width", text="Width")
        self.signal_tree.heading("radix", text="Table radix")
        self.signal_tree.column("signal", width=620, anchor="w")
        self.signal_tree.column("type", width=95, anchor="w", stretch=False)
        self.signal_tree.column("width", width=65, anchor="center", stretch=False)
        self.signal_tree.column("radix", width=140, anchor="w", stretch=False)
        scroll = ttk.Scrollbar(sig_frame, orient="vertical", command=self.signal_tree.yview)
        self.signal_tree.configure(yscrollcommand=scroll.set)
        self.signal_tree.grid(row=1, column=0, sticky="nsew")
        scroll.grid(row=1, column=1, sticky="ns")
        self.signal_tree.bind("<Button-3>", self._show_signal_context_menu)
        self.signal_tree.bind("<Control-Button-1>", self._show_signal_context_menu)

        self.signal_context_menu = tk.Menu(self.root, tearoff=False)
        self.signal_radix_menu = tk.Menu(self.signal_context_menu, tearoff=False)
        self.signal_context_menu.add_cascade(label="Radix", menu=self.signal_radix_menu)
        for radix_key in RADIX_CHOICES:
            label = "Default / WCFG" if radix_key == "auto" else RADIX_LABELS[radix_key]
            self.signal_radix_menu.add_command(label=label, command=lambda r=radix_key: self._set_context_radix(r))
        self.signal_context_menu.add_separator()
        self.signal_context_menu.add_command(label="Remove from selected signals", command=self._remove_highlighted)

        radix_bar = ttk.Frame(sig_frame)
        radix_bar.grid(row=2, column=0, columnspan=2, sticky="ew", pady=(7, 0))
        ttk.Label(radix_bar, text="Default radix:").pack(side="left")
        self.default_radix_combo = ttk.Combobox(radix_bar, state="readonly", width=16, textvariable=self.value_fmt,
                                                 values=[RADIX_LABELS[x] for x in RADIX_CHOICES])
        # Combobox displays labels; normalize in handler.
        self.value_fmt.set(RADIX_LABELS["auto"])
        self.default_radix_combo.pack(side="left", padx=(5, 14))
        self._readonly_controls.append(self.default_radix_combo)
        self.default_radix_combo.bind("<<ComboboxSelected>>", lambda _e: self._refresh_selected_table())
        ttk.Label(radix_bar, text="Right-click signal(s) to set radix.", foreground="gray").pack(side="left")
        ttk.Label(radix_bar, text="Radix affects CSV, JSON and Excel only; VCD has no standard radix metadata.", foreground="gray").pack(side="right")

        export = ttk.LabelFrame(body, text="Export", padding=10)
        body.add(export, weight=1)
        export.columnconfigure(1, weight=1)
        ttk.Label(export, text="Format:").grid(row=0, column=0, sticky="w", pady=4)
        self.format_combo = ttk.Combobox(export, state="readonly", textvariable=self.format_var, width=12,
                                          values=("VCD", "CSV", "JSON", "Excel"))
        self.format_var.set("VCD")
        self.format_combo.grid(row=0, column=1, sticky="w", pady=4)
        self._readonly_controls.append(self.format_combo)
        self.format_combo.bind("<<ComboboxSelected>>", lambda _e: self._format_changed())
        ttk.Label(export, text="Time unit:").grid(row=0, column=2, sticky="e", padx=(20, 5))
        self.time_combo = ttk.Combobox(export, state="readonly", textvariable=self.time_unit, width=7,
                                      values=("fs", "ps", "ns", "us", "ms"))
        self.time_combo.grid(row=0, column=3, sticky="w")
        self._readonly_controls.append(self.time_combo)
        self.json_check = ttk.Checkbutton(export, text="Compact JSON events", variable=self.json_events)
        self.json_check.grid(row=0, column=4, padx=(20, 0), sticky="w")
        self._normal_state_controls.append(self.json_check)

        self.output_label = ttk.Label(export, text="Output:")
        self.output_label.grid(row=1, column=0, sticky="w", pady=4)
        self.output_entry = ttk.Entry(export, textvariable=self.output_file)
        self.output_entry.grid(row=1, column=1, columnspan=3, sticky="ew", pady=4)
        self._normal_state_controls.append(self.output_entry)
        b = ttk.Button(export, text="Browse…", command=self._browse_output); b.grid(row=1, column=4, padx=(8, 0)); self._controls.append(b)
        b = ttk.Button(export, text="Convert", style="Action.TButton", command=self._convert); b.grid(row=2, column=0, columnspan=5, pady=(10, 2)); self._controls.append(b)

        status = ttk.Frame(outer)
        status.grid(row=2, column=0, sticky="ew", pady=(10, 0))
        status.columnconfigure(0, weight=1)
        self.status = ttk.Label(status, text="Ready", foreground="gray")
        self.status.grid(row=0, column=0, sticky="w")
        self.progress = ttk.Progressbar(status, mode="determinate", maximum=100, value=0, length=190)
        self.progress.grid(row=0, column=1, sticky="e")

    @staticmethod
    def _label_to_radix(label: str) -> str:
        for key, text in RADIX_LABELS.items():
            if text == label:
                return key
        return label.lower()

    @staticmethod
    def _format_key(label: str) -> str:
        return {"CSV": "csv", "JSON": "json", "Excel": "excel", "VCD": "vcd"}.get(label, label.lower())

    def _set_busy(self, busy: bool, text: str | None = None) -> None:
        self.busy = busy
        for ctl in self._controls:
            try:
                ctl.configure(state="disabled" if busy else "normal")
            except tk.TclError:
                pass
        for ctl in self._readonly_controls:
            try:
                ctl.configure(state="disabled" if busy else "readonly")
            except tk.TclError:
                pass
        for ctl in self._normal_state_controls:
            try:
                ctl.configure(state="disabled" if busy else "normal")
            except tk.TclError:
                pass
        if not busy:
            fmt0 = self._format_key(self.format_var.get())
            table_enabled = fmt0 != "vcd"
            self.default_radix_combo.configure(state="readonly" if table_enabled else "disabled")
            self.time_combo.configure(state="readonly" if table_enabled else "disabled")
            self.json_check.configure(state="normal" if fmt0 == "json" else "disabled")
        if busy:
            self.progress.stop()
            self.progress.configure(mode="indeterminate", maximum=100, value=0)
            self.progress.start(10)
        else:
            self.progress.stop()
            self.progress.configure(mode="determinate", maximum=100, value=0)
        if text:
            self.status.config(text=text, foreground="blue" if busy else "gray")
        self.root.update_idletasks()

    def _run_background(self, tag: str, func) -> None:
        if self.busy:
            return
        self._set_busy(True)
        def worker() -> None:
            try:
                result = func()
                self._task_queue.put((tag, True, result))
            except BaseException as exc:
                self._task_queue.put((tag, False, exc))
        threading.Thread(target=worker, name=f"waveform-{tag}", daemon=True).start()
        self.root.after(75, self._poll_background)

    def _poll_background(self) -> None:
        try:
            tag, ok, payload = self._task_queue.get_nowait()
        except queue.Empty:
            if self.busy:
                self.root.after(75, self._poll_background)
            return
        if tag == "batch-progress":
            done, total, name = payload
            self.progress.stop()
            self.progress.configure(mode="determinate", maximum=max(1, total), value=done)
            self.status.config(text=f"Batch {done}/{total}: {name}", foreground="blue")
            if self.busy:
                self.root.after(25, self._poll_background)
            return
        self._set_busy(False)
        if not ok:
            self.status.config(text=f"Operation failed: {payload}", foreground="red")
            if not (tag == "load" and self._current_load_quiet):
                messagebox.showerror("Waveform Converter", str(payload), parent=self.root)
            self._pending_after_load = None
            if tag == "load":
                self._loading_signature = None
                self._current_load_quiet = False
            if self._reload_after_busy:
                self._reload_after_busy = False
                self.root.after(50, self._auto_load_if_ready)
            return
        if tag == "load":
            self._current_load_quiet = False
            self._finish_load(payload)
        elif tag == "export":
            self._finish_export(payload)
        elif tag == "batch-export":
            self._finish_batch_export(payload)
        if self._reload_after_busy:
            self._reload_after_busy = False
            self.root.after(50, self._auto_load_if_ready)

    def _source_signature(self) -> tuple[str, str]:
        in_paths = self.input_paths or ([Path(self.input_file.get().strip())] if self.input_file.get().strip() else [])
        cfg_paths = self.wcfg_paths or ([Path(self.wcfg_file.get().strip())] if self.wcfg_file.get().strip() else [])
        return ("|".join(str(p) for p in in_paths), "|".join(str(p) for p in cfg_paths))

    def _input_text_changed(self, *_args) -> None:
        if not self._suppress_auto_load:
            text = self.input_file.get().strip()
            if len(self.input_paths) != 1 or (self.input_paths and str(self.input_paths[0]) != text):
                self.input_paths = [Path(text)] if text else []
        self._source_text_changed()

    def _wcfg_text_changed(self, *_args) -> None:
        if not self._suppress_auto_load:
            text = self.wcfg_file.get().strip()
            if len(self.wcfg_paths) != 1 or (self.wcfg_paths and str(self.wcfg_paths[0]) != text):
                self.wcfg_paths = [Path(text)] if text else []
        self._source_text_changed()

    def _source_text_changed(self, *_args) -> None:
        if self._suppress_auto_load:
            return
        self._invalidate_loaded()
        self._update_source_hint()
        if self._auto_load_after_id is not None:
            try:
                self.root.after_cancel(self._auto_load_after_id)
            except tk.TclError:
                pass
        self._auto_load_after_id = self.root.after(450, self._auto_load_if_ready)

    def _auto_load_if_ready(self) -> None:
        self._auto_load_after_id = None
        source_paths = self.input_paths or ([Path(self.input_file.get().strip())] if self.input_file.get().strip() else [])
        cfg_paths = self.wcfg_paths or ([Path(self.wcfg_file.get().strip())] if self.wcfg_file.get().strip() else [])
        if not source_paths:
            return
        if any(p.suffix.lower() not in {".wdb", ".wcfg", ".vcd"} or not p.is_file() for p in source_paths):
            return
        if any(p.suffix.lower() != ".wcfg" or not p.is_file() for p in cfg_paths):
            return
        try:
            jobs = build_batch_jobs(source_paths, cfg_paths)
        except (ConversionError, DecodeError):
            return
        first = jobs[0]
        sig = ("|".join(str(p.resolve()) for p in first.inputs), str(first.explicit_wcfg.resolve()) if first.explicit_wcfg else "")
        if sig == self._loaded_signature or sig == self._loading_signature:
            return
        if self.busy:
            self._reload_after_busy = True
            return
        self._begin_load(None, quiet=True)

    def _set_default_format_for_path(self, path: Path) -> None:
        self.format_var.set("CSV" if path.suffix.lower() == ".vcd" else "VCD")

    def _browse_input(self) -> None:
        filenames = filedialog.askopenfilenames(
            title="Select waveform",
            filetypes=[("Waveform files", "*.wdb *.wcfg *.vcd"), ("Vivado WDB", "*.wdb"),
                       ("Vivado WCFG", "*.wcfg"), ("VCD", "*.vcd"), ("All files", "*.*")])
        if not filenames:
            return
        paths = [Path(x).resolve() for x in filenames]
        self._suppress_auto_load = True
        try:
            self.input_paths = paths
            self.input_file.set(str(paths[0]))
            # WCFG-only primary input does not need a separate WCFG field.
            if all(p.suffix.lower() == ".wcfg" for p in paths):
                self.wcfg_paths = []
                self.wcfg_file.set("")
        finally:
            self._suppress_auto_load = False
        self._invalidate_loaded()
        self._set_default_format_for_path(paths[0])
        self._update_source_hint()
        self._update_output_guess()
        self._begin_load(None)

    def _browse_wcfg(self) -> None:
        filenames = filedialog.askopenfilenames(title="Select WCFG", filetypes=[("Vivado WCFG", "*.wcfg"), ("All files", "*.*")])
        if filenames:
            paths = [Path(x).resolve() for x in filenames]
            self._suppress_auto_load = True
            try:
                self.wcfg_paths = paths
                self.wcfg_file.set(str(paths[0]))
            finally:
                self._suppress_auto_load = False
            self._invalidate_loaded()
            self._update_source_hint()
            if self.input_file.get().strip():
                self._begin_load(None)

    def _update_source_hint(self) -> None:
        nsrc = len(self.input_paths) if self.input_paths else (1 if self.input_file.get().strip() else 0)
        ncfg = len(self.wcfg_paths) if self.wcfg_paths else (1 if self.wcfg_file.get().strip() else 0)
        if nsrc > 1 or ncfg > 1:
            self.source_hint.config(text=f"Batch mode: {nsrc} waveform selection(s), {ncfg} WCFG selection(s). Matching WCFG files are paired automatically.")
        else:
            self.source_hint.config(text="Sources load automatically. Browse supports multi-select; matching WCFG files are paired automatically.")

    def _invalidate_loaded(self) -> None:
        self.wf = None
        self._loaded_signature = None
        self.batch_jobs = []
        self.selected_names.clear()
        self.radix_overrides.clear()
        self.signal_tree.delete(*self.signal_tree.get_children())
        self.signal_summary.config(text="Not loaded", foreground="gray")
        self.status.config(text="Ready", foreground="gray")

    def _begin_load(self, after: str | None, quiet: bool = False) -> None:
        source_paths = self.input_paths or ([Path(self.input_file.get().strip())] if self.input_file.get().strip() else [])
        cfg_paths = self.wcfg_paths or ([Path(self.wcfg_file.get().strip())] if self.wcfg_file.get().strip() else [])
        if not source_paths:
            if not quiet:
                messagebox.showwarning("Waveform Converter", "Select a waveform file first.", parent=self.root)
            return
        try:
            jobs = build_batch_jobs(source_paths, cfg_paths)
        except (ConversionError, DecodeError) as exc:
            if not quiet:
                messagebox.showerror("Waveform Converter", str(exc), parent=self.root)
            return
        self.batch_jobs = jobs
        job = jobs[0]
        if self.busy:
            self._reload_after_busy = True
            return
        self._pending_after_load = after
        self._current_load_quiet = quiet
        self._loading_signature = ("|".join(str(p.resolve()) for p in job.inputs), str(job.explicit_wcfg.resolve()) if job.explicit_wcfg else "")
        suffix = f" (preview 1/{len(jobs)})" if len(jobs) > 1 else ""
        self.status.config(text=f"Decoding waveform{suffix}…", foreground="blue")
        self._run_background("load", lambda: load_waveform(list(job.inputs), job.explicit_wcfg))

    def _finish_load(self, wf: Waveform) -> None:
        self.wf = wf
        self._loaded_signature = self._loading_signature
        self._loading_signature = None
        self.selected_names = default_gui_selection(wf)
        self.radix_overrides.clear()
        # WDB/WCFG conversion is primarily a WDB -> VCD workflow. Ordinary
        # VCD input defaults to CSV because VCD -> VCD is intentionally absent.
        self.format_var.set("CSV" if wf.source_kind == "vcd" else "VCD")
        self._update_format_options()
        self._refresh_selected_table()
        self._update_output_guess()
        detail = f"{len(wf.streams)} runtime streams, {len(wf.signal_order)} declarations"
        if wf.default_selected is not None:
            detail += f"; WCFG selected {len(wf.default_selected)}"
        if len(self.batch_jobs) > 1:
            detail += f"; batch {len(self.batch_jobs)} files"
        self.status.config(text=f"Loaded {wf.source.name}: {detail}", foreground="green")
        action = self._pending_after_load
        self._pending_after_load = None
        if action == "select":
            self.root.after_idle(self._select_signals)
        elif action == "convert":
            self.root.after_idle(self._start_export)

    def _update_format_options(self) -> None:
        if self.wf and self.wf.source_kind == "vcd":
            self.format_combo.configure(values=("CSV", "JSON", "Excel"))
            if self._format_key(self.format_var.get()) == "vcd":
                self.format_var.set("CSV")
        else:
            self.format_combo.configure(values=("VCD", "CSV", "JSON", "Excel"))
        self._format_changed()

    def _select_signals(self) -> None:
        if self.busy:
            return
        if self.wf is None:
            self._begin_load("select")
            return
        dlg = SignalSelectorDialog(self.root, self.wf, self.selected_names)
        if dlg.result is not None:
            self.selected_names = dlg.result
            # Drop stale overrides when a signal is removed; re-adding gets WCFG/default radix.
            self.radix_overrides = {k: v for k, v in self.radix_overrides.items() if k in self.selected_names}
            self._refresh_selected_table()

    def _remove_highlighted(self) -> None:
        if not self.wf:
            return
        for key in self.signal_tree.selection():
            if key in self.wf.signals:
                self.selected_names.discard(self.wf.signals[key].name)
                self.radix_overrides.pop(self.wf.signals[key].name, None)
        if not self.selected_names:
            messagebox.showwarning("Waveform Converter", "At least one signal must remain selected.", parent=self.root)
            self.selected_names = default_gui_selection(self.wf)
        self._refresh_selected_table()

    def _refresh_selected_table(self) -> None:
        self.signal_tree.delete(*self.signal_tree.get_children())
        if not self.wf:
            self.signal_summary.config(text="No waveform loaded", foreground="gray")
            return
        default_fmt = self._label_to_radix(self.value_fmt.get())
        selected_count = 0
        selected_streams_set: set[str] = set()
        for key in self.wf.signal_order:
            sig = self.wf.signals[key]
            if sig.name not in self.selected_names:
                continue
            selected_count += 1
            selected_streams_set.add(sig.stream_key)
            override = self.radix_overrides.get(sig.name)
            rules = [(sig.name, override)] if override else None
            radix = effective_radix(sig, default_fmt, rules)
            label = RADIX_LABELS.get(radix, radix)
            if override:
                label += " (custom)"
            elif default_fmt == "auto" and sig.wcfg_radix:
                label += " (WCFG)"
            typ = "real" if sig.is_real else sig.var_type
            self.signal_tree.insert("", "end", iid=key, values=(sig.name, typ, sig.width, label))
        self.signal_summary.config(text=f"{selected_count} selected / {len(self.wf.signal_order)} available; {len(selected_streams_set)} streams",
                                   foreground="gray")

    def _show_signal_context_menu(self, event) -> str:
        if not self.wf:
            return "break"
        row = self.signal_tree.identify_row(event.y)
        if not row:
            return "break"
        current = set(self.signal_tree.selection())
        if row not in current:
            self.signal_tree.selection_set(row)
        try:
            self.signal_context_menu.tk_popup(event.x_root, event.y_root)
        finally:
            self.signal_context_menu.grab_release()
        return "break"

    def _set_context_radix(self, fmt: str) -> None:
        if not self.wf or fmt not in RADIX_CHOICES:
            return
        for key in self.signal_tree.selection():
            sig = self.wf.signals.get(key)
            if not sig or sig.is_real:
                continue
            if fmt == "auto":
                self.radix_overrides.pop(sig.name, None)
            else:
                self.radix_overrides[sig.name] = fmt
        self._refresh_selected_table()

    def _reset_radix(self) -> None:
        if not self.wf:
            return
        items = self.signal_tree.selection()
        if items:
            for key in items:
                sig = self.wf.signals.get(key)
                if sig:
                    self.radix_overrides.pop(sig.name, None)
        else:
            self.radix_overrides.clear()
        self._refresh_selected_table()

    def _format_changed(self) -> None:
        fmt = self._format_key(self.format_var.get())
        table_enabled = fmt != "vcd"
        if not self.busy:
            self.default_radix_combo.configure(state="readonly" if table_enabled else "disabled")
            self.time_combo.configure(state="readonly" if table_enabled else "disabled")
            self.json_check.configure(state="normal" if fmt == "json" else "disabled")
        self._update_output_guess()

    def _update_output_guess(self) -> None:
        p = self.input_file.get().strip()
        if not p:
            return
        fmt = self._format_key(self.format_var.get())
        if fmt not in {"vcd", "csv", "json", "excel"}:
            return
        if len(self.batch_jobs) > 1 or len(self.input_paths) > 1 or len(self.wcfg_paths) > 1:
            self.output_label.config(text="Output folder:")
            self.output_file.set(str(ensure_output_dir()))
            return
        self.output_label.config(text="Output:")
        source = self.wf.source if self.wf is not None else Path(p)
        self.output_file.set(str(default_output_path(source, fmt)))

    def _browse_output(self) -> None:
        if len(self.batch_jobs) > 1 or len(self.input_paths) > 1 or len(self.wcfg_paths) > 1:
            folder = filedialog.askdirectory(title="Select batch output folder", initialdir=str(ensure_output_dir()))
            if folder:
                self.output_file.set(folder)
            return
        fmt = self._format_key(self.format_var.get())
        ext = {"vcd": ".vcd", "csv": ".csv", "json": ".json", "excel": ".xlsx"}[fmt]
        filename = filedialog.asksaveasfilename(title="Save converted waveform", defaultextension=ext,
                                                initialdir=str(ensure_output_dir()),
                                                filetypes=[(fmt.upper(), f"*{ext}"), ("All files", "*.*")])
        if filename:
            self.output_file.set(filename)

    def _convert(self) -> None:
        if self.busy:
            return
        if self.wf is None:
            self._begin_load("convert")
            return
        self._start_export()

    def _start_export(self) -> None:
        if not self.wf:
            return
        if not self.selected_names:
            messagebox.showwarning("Waveform Converter", "Select at least one signal.", parent=self.root)
            return
        fmt = self._format_key(self.format_var.get())
        if self.wf.source_kind == "vcd" and fmt == "vcd":
            messagebox.showwarning("Waveform Converter", "VCD to VCD conversion is not required; choose CSV, JSON or Excel.", parent=self.root)
            return
        if len(self.batch_jobs) > 1:
            self._start_batch_export(fmt)
            return
        output = Path(self.output_file.get().strip()) if self.output_file.get().strip() else default_output_path(self.wf.source, fmt)
        default_fmt = self._label_to_radix(self.value_fmt.get())
        radix_rules = [(name, fmt0) for name, fmt0 in self.radix_overrides.items()]
        wf2 = filter_waveform(self.wf, selected=set(self.selected_names))
        json_events = self.json_events.get()
        time_unit = self.time_unit.get()
        self.status.config(text=f"Writing {fmt.upper()}…", foreground="blue")

        def task():
            count = export_waveform(wf2, output, fmt, time_unit, default_fmt,
                                    json_events=json_events, radix_rules=radix_rules)
            log_path = write_wdb_log_if_needed(self.wf, output)
            return output, count, log_path, fmt
        self._run_background("export", task)

    def _finish_export(self, result) -> None:
        output, count, log_path, fmt = result
        log_note = f"; log {log_path.name}" if log_path else ""
        self.status.config(text=f"Saved {count:,} rows/events → {output}{log_note}", foreground="green")

    def _start_batch_export(self, fmt: str) -> None:
        jobs = list(self.batch_jobs)
        if not jobs:
            return
        outdir_text = self.output_file.get().strip()
        outdir = Path(outdir_text) if outdir_text else ensure_output_dir()
        outdir.mkdir(parents=True, exist_ok=True)
        default_fmt = self._label_to_radix(self.value_fmt.get())
        radix_rules = [(name, fmt0) for name, fmt0 in self.radix_overrides.items()]
        json_events = self.json_events.get()
        time_unit = self.time_unit.get()
        preview_source = self.wf.source.resolve() if self.wf else None
        preview_selected = set(self.selected_names)
        self.status.config(text=f"Batch converting {len(jobs)} files…", foreground="blue")

        def task() -> BatchExportResult:
            successes: list[tuple[Path, int, Path | None]] = []
            failures: list[tuple[Path, str]] = []
            used: set[Path] = set()
            for idx, job in enumerate(jobs, 1):
                self._task_queue.put(("batch-progress", True, (idx - 1, len(jobs), job.source_hint.name)))
                try:
                    wf = load_waveform(list(job.inputs), job.explicit_wcfg)
                    if wf.source_kind == "vcd" and fmt == "vcd":
                        raise ConversionError("VCD to VCD conversion is intentionally not supported")
                    selected = default_gui_selection(wf)
                    if preview_source is not None and wf.source.resolve() == preview_source:
                        selected = preview_selected
                    wf2 = filter_waveform(wf, selected=set(selected))
                    output = unique_batch_output(outdir.resolve(), wf.source, fmt, used)
                    count = export_waveform(wf2, output, fmt, time_unit, default_fmt,
                                            json_events=json_events, radix_rules=radix_rules)
                    log_path = write_wdb_log_if_needed(wf, output)
                    successes.append((output, count, log_path))
                except BaseException as exc:
                    failures.append((job.source_hint, f"{type(exc).__name__}: {exc}"))
                self._task_queue.put(("batch-progress", True, (idx, len(jobs), job.source_hint.name)))

            report = outdir.resolve() / "batch_conversion_report.txt"
            lines = [
                f"Vivado Waveform Converter {APP_VERSION}",
                f"Format: {fmt.upper()}",
                f"Jobs: {len(jobs)}",
                f"Succeeded: {len(successes)}",
                f"Failed: {len(failures)}",
                "",
                "SUCCESS",
            ]
            for output, count, log_path in successes:
                lines.append(f"{output} | {count} rows/events" + (f" | log={log_path}" if log_path else ""))
            lines.extend(["", "FAILURES"])
            for source, error in failures:
                lines.append(f"{source} | {error}")
            report.write_text("\n".join(lines) + "\n", encoding="utf-8")
            return BatchExportResult(successes, failures, report, fmt)

        self._run_background("batch-export", task)

    def _finish_batch_export(self, result: BatchExportResult) -> None:
        if result.failures:
            self.status.config(
                text=f"Batch finished: {len(result.successes)} succeeded, {len(result.failures)} failed — {result.report_path}",
                foreground="red",
            )
            details = "\n".join(f"{src.name}: {err}" for src, err in result.failures[:8])
            if len(result.failures) > 8:
                details += f"\n… and {len(result.failures) - 8} more"
            messagebox.showwarning(
                "Waveform Converter",
                f"Batch completed with failures.\n\n{details}\n\nReport:\n{result.report_path}",
                parent=self.root,
            )
        else:
            total_rows = sum(count for _output, count, _log in result.successes)
            self.status.config(
                text=f"Batch complete: {len(result.successes)} files, {total_rows:,} rows/events → {result.report_path.parent}",
                foreground="green",
            )

    def run(self) -> None:
        self.root.mainloop()


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def normalize_radix_name(value: str) -> str:
    v = value.strip().lower().replace("_", "-")
    aliases = {
        "auto": "auto", "wcfg": "auto",
        "hex": "hex", "hexadecimal": "hex",
        "int": "int", "uint": "int", "unsigned": "int", "decimal": "int", "unsigned-decimal": "int",
        "signed": "signed", "signed-decimal": "signed",
        "bin": "bin", "binary": "bin",
        "oct": "oct", "octal": "oct",
        "smag": "smag", "sign-magnitude": "smag", "signed-magnitude": "smag",
    }
    if v not in aliases:
        raise ConversionError(f"unsupported radix {value!r}; use auto, hex, unsigned, signed, bin, oct or smag")
    return aliases[v]


def parse_radix_rules(values: list[str]) -> list[tuple[str, str]]:
    rules: list[tuple[str, str]] = []
    for item in values:
        if "=" not in item:
            raise ConversionError(f"invalid --radix {item!r}; expected PATTERN=RADIX")
        pattern, radix = item.rsplit("=", 1)
        pattern = pattern.strip()
        if not pattern:
            raise ConversionError("--radix pattern cannot be empty")
        rules.append((pattern, normalize_radix_name(radix)))
    return rules


def build_arg_parser() -> argparse.ArgumentParser:
    ap = argparse.ArgumentParser(
        description="Convert Vivado/XSim WDB/WCFG or VCD waveform data to VCD/table formats.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""Examples:\n  waveform_converter.py sim.wdb --vcd\n  waveform_converter.py sim.wdb sim.wcfg --csv\n  waveform_converter.py a.wdb a.wcfg b.wdb b.wcfg --vcd\n  waveform_converter.py sim.wcfg --excel\n  waveform_converter.py waveform.vcd --json\n  waveform_converter.py sim.wdb --csv --radix '*counter*=unsigned' --radix '*signed*=signed'\n""")
    ap.add_argument("inputs", nargs="*", type=Path, help="one or more VCD/WDB/WCFG inputs; multiple WDB/WCFG files form a batch")
    ap.add_argument("--version", action="version", version=f"%(prog)s {APP_VERSION}")
    ap.add_argument("--wcfg", type=Path, help="optional WCFG for a WDB input")
    ap.add_argument("-o", "--output", type=Path)
    ap.add_argument("--format", choices=["vcd", "csv", "json", "excel"])
    ap.add_argument("--vcd", action="store_true", help="WDB/WCFG input only")
    ap.add_argument("--csv", action="store_true")
    ap.add_argument("--json", action="store_true")
    ap.add_argument("--excel", action="store_true")
    ap.add_argument("--auto", dest="value_fmt", action="store_const", const="auto", help="use WCFG radix when available (default)")
    ap.add_argument("--hex", dest="value_fmt", action="store_const", const="hex")
    ap.add_argument("--int", "--unsigned", dest="value_fmt", action="store_const", const="int")
    ap.add_argument("--signed", dest="value_fmt", action="store_const", const="signed")
    ap.add_argument("--smag", dest="value_fmt", action="store_const", const="smag")
    ap.add_argument("--bin", dest="value_fmt", action="store_const", const="bin")
    ap.add_argument("--oct", dest="value_fmt", action="store_const", const="oct")
    ap.set_defaults(value_fmt="auto")
    for unit in ["fs", "ps", "ns", "us", "ms"]:
        ap.add_argument(f"--{unit}", dest="time_unit", action="store_const", const=unit)
    ap.set_defaults(time_unit="us")
    ap.add_argument("--timescale", help="WDB->VCD output timescale, e.g. 1ps or 1ns")
    ap.add_argument("--include", action="append", default=[])
    ap.add_argument("--exclude", action="append", default=[])
    ap.add_argument("--primary-only", action="store_true", help="one declaration per alias/runtime stream")
    ap.add_argument("--all-stored", action="store_true", help="ignore WCFG selection and use all waveforms stored in WDB")
    ap.add_argument("--signals", action="store_true", help="list available declarations and exit")
    ap.add_argument("--radix", action="append", default=[], metavar="PATTERN=RADIX",
                    help="per-signal table radix; repeatable; glob patterns accepted")
    ap.add_argument("--json-events", action="store_true", help="compact event-oriented JSON")
    ap.add_argument("--gui", action="store_true")
    return ap


def cli_main(argv: list[str] | None = None) -> int:
    ns = build_arg_parser().parse_args(argv)
    if ns.gui or not ns.inputs:
        if not HAS_TK:
            raise ConversionError("GUI requires tkinter")
        ConverterGUI().run()
        return 0

    chosen = [name for name in ["vcd", "csv", "json", "excel"] if getattr(ns, name)]
    if ns.format and chosen:
        raise ConversionError("use either --format or one of --vcd/--csv/--json/--excel")
    if len(chosen) > 1:
        raise ConversionError("choose only one output format")
    explicit_fmt = ns.format or (chosen[0] if chosen else None)

    jobs = build_batch_jobs(ns.inputs, [ns.wcfg] if ns.wcfg else [])

    if len(jobs) > 1:
        source_kind = "vcd" if all(j.source_hint.suffix.lower() == ".vcd" for j in jobs) else "wdb"
        fmt = explicit_fmt or ("vcd" if source_kind == "wdb" else "csv")
        if source_kind == "vcd" and fmt == "vcd":
            raise ConversionError("VCD to VCD conversion is intentionally not supported")
        outdir = (ns.output or ensure_output_dir()).resolve()
        if ns.output and ns.output.suffix and not ns.output.is_dir():
            raise ConversionError("--output must be a directory when converting a batch")
        outdir.mkdir(parents=True, exist_ok=True)
        radix_rules = parse_radix_rules(ns.radix)
        used: set[Path] = set()
        failures: list[tuple[Path, str]] = []
        total = 0
        print(f"Batch  : {len(jobs)} inputs -> {fmt.upper()}")
        print(f"Output : {outdir}")
        for i, job in enumerate(jobs, 1):
            try:
                wf = load_waveform(list(job.inputs), job.explicit_wcfg)
                print(f"[{i}/{len(jobs)}] {wf.source}")
                if ns.signals:
                    print_signals(wf)
                    continue
                selected = None if ns.all_stored else wf.default_selected
                wf = filter_waveform(wf, ns.include or None, ns.exclude or None,
                                     selected=(set(selected) if selected is not None else None),
                                     primary_only=ns.primary_only)
                output = unique_batch_output(outdir, wf.source, fmt, used)
                count = export_waveform(wf, output, fmt, ns.time_unit, ns.value_fmt,
                                        ns.timescale, ns.json_events, radix_rules)
                write_wdb_log_if_needed(wf, output)
                total += count
                print(f"        -> {output.name} ({count:,} rows/events)")
            except (ConversionError, DecodeError, OSError, ValueError, struct.error) as exc:
                failures.append((job.source_hint, str(exc)))
                print(f"        ERROR: {exc}", file=sys.stderr)
        if ns.signals:
            return 0
        report = outdir / "batch_conversion_report.txt"
        lines = [f"Vivado Waveform Converter {APP_VERSION}", f"Jobs: {len(jobs)}", f"Failed: {len(failures)}", ""]
        lines.extend(f"FAIL {src}: {err}" for src, err in failures)
        report.write_text("\n".join(lines) + "\n", encoding="utf-8")
        print(f"Done   : {len(jobs)-len(failures)}/{len(jobs)} files, {total:,} rows/events")
        print(f"Report : {report}")
        return 3 if failures else 0

    job = jobs[0]
    wf = load_waveform(list(job.inputs), job.explicit_wcfg)
    # Match the GUI and drag-and-drop workflow: existing WDB/WCFG input is
    # primarily converted to VCD, while ordinary VCD input defaults to CSV.
    fmt = explicit_fmt or ("vcd" if wf.source_kind == "wdb" else "csv")
    if wf.source_kind == "vcd" and fmt == "vcd":
        raise ConversionError("VCD to VCD conversion is intentionally not supported")

    print(f"Input  : {wf.source} ({wf.source_kind.upper()})")
    print(f"Signals: {len(wf.signal_order)} declarations across {len(wf.streams)} runtime streams")
    if wf.missing_wcfg:
        print(f"WARNING: {len(wf.missing_wcfg)} WCFG object(s) are not stored in the WDB")

    if ns.signals:
        print_signals(wf)
        return 0

    selected = None if ns.all_stored else wf.default_selected
    wf = filter_waveform(wf, ns.include or None, ns.exclude or None,
                         selected=(set(selected) if selected is not None else None),
                         primary_only=ns.primary_only)
    radix_rules = parse_radix_rules(ns.radix)
    output = (ns.output or default_output_path(wf.source, fmt)).resolve()
    print(f"Output : {output} ({fmt.upper()})")
    count = export_waveform(wf, output, fmt, ns.time_unit, ns.value_fmt,
                            ns.timescale, ns.json_events, radix_rules)
    log_path = write_wdb_log_if_needed(wf, output)
    print(f"Done   : {count:,} rows/events")
    if log_path:
        print(f"Log    : {log_path}")
    return 0


def main() -> None:
    try:
        raise SystemExit(cli_main())
    except (ConversionError, DecodeError, OSError, ValueError, struct.error) as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        raise SystemExit(3)


if __name__ == "__main__":
    main()
